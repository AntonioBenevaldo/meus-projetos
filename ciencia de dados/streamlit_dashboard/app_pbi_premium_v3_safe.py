# app_pbi_premium_v3_safe.py
# Dashboard Executivo (Power BI Premium v3 SAFE) - Streamlit
# Autor: Benevaldo + ChatGPT
# Requisitos: streamlit pandas numpy plotly openpyxl

from __future__ import annotations

import re
from io import BytesIO
from datetime import datetime, timedelta, date
from typing import Dict, Tuple, Optional, List

import numpy as np
import pandas as pd
import streamlit as st
import plotly.express as px


# =========================
# CONFIG STREAMLIT
# =========================
st.set_page_config(
    page_title="Dashboard Executivo (Power BI Premium v3)",
    layout="wide",
    initial_sidebar_state="collapsed",  # menu recolhível por padrão
)


# =========================
# CSS (Power BI Dark)
# =========================
def inject_css() -> None:
    st.markdown(
        """
        <style>
        :root{
          --bg:#0f1117;
          --panel:#141824;
          --card:#161b2b;
          --muted:#9aa4b2;
          --text:#e6edf3;
          --line:#24314a;
          --accent:#4aa3ff;
          --danger:#ff5b5b;
        }

        /* fundo geral */
        .stApp { background: var(--bg) !important; color: var(--text) !important; }

        /* sidebar */
        section[data-testid="stSidebar"] {
            background: linear-gradient(180deg, #0b0e16, #0f1117) !important;
            border-right: 1px solid var(--line);
        }

        /* blocos */
        div[data-testid="stVerticalBlock"] > div:has(> div[data-testid="stMetric"]) {
            background: var(--card);
            border: 1px solid rgba(36,49,74,.55);
            padding: 14px 14px 6px 14px;
            border-radius: 14px;
        }

        /* cards gerais */
        .pbi-card{
            background: var(--card);
            border: 1px solid rgba(36,49,74,.55);
            border-radius: 16px;
            padding: 16px;
        }

        /* título */
        .pbi-title{
            font-size: 34px;
            font-weight: 800;
            letter-spacing: .2px;
            margin-bottom: 2px;
        }
        .pbi-sub{
            color: var(--muted);
            font-size: 13px;
            margin-top: -4px;
            margin-bottom: 16px;
        }

        /* botão */
        .stButton>button {
            border-radius: 12px !important;
            border: 1px solid rgba(36,49,74,.7) !important;
            background: rgba(22,27,43,.8) !important;
            color: var(--text) !important;
        }
        .stButton>button:hover{
            border-color: rgba(74,163,255,.9) !important;
            box-shadow: 0 0 0 2px rgba(74,163,255,.15);
        }

        /* inputs */
        div[data-baseweb="input"] input,
        div[data-baseweb="select"] > div,
        textarea {
            background: rgba(22,27,43,.8) !important;
            border-radius: 12px !important;
            border-color: rgba(36,49,74,.75) !important;
            color: var(--text) !important;
        }

        /* tabelas */
        .stDataFrame {
            border-radius: 14px !important;
            overflow: hidden !important;
            border: 1px solid rgba(36,49,74,.55) !important;
        }

        /* tabs */
        button[role="tab"]{
            border-radius: 12px !important;
        }

        /* links */
        a{ color: var(--accent) !important; }

        </style>
        """,
        unsafe_allow_html=True,
    )


inject_css()


# =========================
# HELPERS - DADOS
# =========================
def _std_col_name(c: str) -> str:
    c = str(c).strip().lower()
    c = re.sub(r"[^\w\s]", "", c, flags=re.UNICODE)
    c = c.replace(" ", "_")
    c = re.sub(r"_+", "_", c)
    return c


def _pick_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    cols = list(df.columns)
    for c in candidates:
        if c in cols:
            return c
    return None


def _parse_brl_money_to_float(series: pd.Series) -> pd.Series:
    """
    Converte strings tipo: 'R$ 12.781.406,01' -> 12781406.01
    Aceita também números puros.
    """
    if series.dtype.kind in "if":
        return series.astype(float)

    s = series.astype(str).str.strip()

    # remove moeda e espaços
    s = s.str.replace("R$", "", regex=False).str.replace("r$", "", regex=False)
    s = s.str.replace("\u00a0", " ", regex=False).str.replace(" ", "", regex=False)

    # remove separador de milhar '.' e troca decimal ',' por '.'
    s = s.str.replace(".", "", regex=False)
    s = s.str.replace(",", ".", regex=False)

    # mantém apenas dígitos, sinal e ponto
    s = s.str.replace(r"[^0-9\.\-]", "", regex=True)

    return pd.to_numeric(s, errors="coerce")


def normalize_dataset(df_raw: pd.DataFrame) -> pd.DataFrame:
    """
    Normaliza nomes de colunas e garante colunas essenciais:
    data, categoria, marca, produto, uf, canal, quantidade, preco_unitario, valor_total
    """
    df = df_raw.copy()

    # padroniza nomes
    df.columns = [_std_col_name(c) for c in df.columns]

    # mapa de possíveis nomes
    map_date = ["data", "dt", "date", "dia", "data_venda"]
    map_cat = ["categoria", "category", "grupo", "familia"]
    map_brand = ["marca", "brand", "fabricante"]
    map_prod = ["produto", "product", "descricao", "item"]
    map_uf = ["uf", "estado", "state"]
    map_channel = ["canal", "channel", "origem", "plataforma"]
    map_qty = ["quantidade", "qtd", "qtde", "quantity", "unidades"]
    map_unit = ["preco_unitario", "preco", "valor_unitario", "unit_price", "price"]
    map_total = ["valor_total", "total", "faturamento", "receita", "valor", "sales"]

    c_data = _pick_col(df, map_date)
    c_cat = _pick_col(df, map_cat)
    c_brand = _pick_col(df, map_brand)
    c_prod = _pick_col(df, map_prod)
    c_uf = _pick_col(df, map_uf)
    c_channel = _pick_col(df, map_channel)
    c_qty = _pick_col(df, map_qty)
    c_unit = _pick_col(df, map_unit)
    c_total = _pick_col(df, map_total)

    # renomeia para padrão
    rename_map = {}
    if c_data: rename_map[c_data] = "data"
    if c_cat: rename_map[c_cat] = "categoria"
    if c_brand: rename_map[c_brand] = "marca"
    if c_prod: rename_map[c_prod] = "produto"
    if c_uf: rename_map[c_uf] = "uf"
    if c_channel: rename_map[c_channel] = "canal"
    if c_qty: rename_map[c_qty] = "quantidade"
    if c_unit: rename_map[c_unit] = "preco_unitario"
    if c_total: rename_map[c_total] = "valor_total"

    df = df.rename(columns=rename_map)

    # colunas mínimas (cria se faltar)
    for col in ["categoria", "marca", "produto", "uf", "canal"]:
        if col not in df.columns:
            df[col] = "N/A"

    # data
    if "data" not in df.columns:
        # cria uma data falsa (para não quebrar)
        df["data"] = pd.Timestamp.today().normalize()
    df["data"] = pd.to_datetime(df["data"], errors="coerce")
    df["data"] = df["data"].fillna(pd.Timestamp.today().normalize())

    # quantidade
    if "quantidade" not in df.columns:
        df["quantidade"] = 1
    df["quantidade"] = pd.to_numeric(df["quantidade"], errors="coerce").fillna(0)

    # preco unitário
    if "preco_unitario" in df.columns:
        df["preco_unitario"] = _parse_brl_money_to_float(df["preco_unitario"]).fillna(0)
    else:
        df["preco_unitario"] = 0.0

    # valor total (se não existir, calcula)
    if "valor_total" in df.columns:
        df["valor_total"] = _parse_brl_money_to_float(df["valor_total"]).fillna(0)
    else:
        df["valor_total"] = df["quantidade"] * df["preco_unitario"]

    # padroniza textos
    for col in ["categoria", "marca", "produto", "uf", "canal"]:
        df[col] = df[col].astype(str).str.strip()
        df[col] = df[col].replace({"": "N/A", "nan": "N/A", "None": "N/A"})

    # derivados
    df["ano"] = df["data"].dt.year
    df["mes"] = df["data"].dt.to_period("M").astype(str)
    df["dia"] = df["data"].dt.date

    return df


@st.cache_data(show_spinner=False)
def demo_data(seed: int = 42) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    datas = pd.date_range("2025-01-01", "2025-12-31", freq="D")

    categorias = ["CELULARES", "TV & VÍDEO", "INFORMÁTICA", "ÁUDIO", "ACESSÓRIOS"]
    marcas = {
        "CELULARES": ["SAMSUNG", "APPLE", "MOTOROLA", "XIAOMI"],
        "TV & VÍDEO": ["SAMSUNG", "LG", "SONY", "PHILIPS"],
        "INFORMÁTICA": ["DELL", "LENOVO", "ACER", "ASUS"],
        "ÁUDIO": ["JBL", "SONY", "BOSE", "PHILIPS"],
        "ACESSÓRIOS": ["GENÉRICO", "MULTILASER", "LOGITECH", "SANDISK"],
    }
    produtos = {}
    for cat in categorias:
        produtos[cat] = {}
        for m in marcas[cat]:
            produtos[cat][m] = [f"{cat} - {m} - ITEM {i}" for i in range(1, 9)]

    ufs = ["SP", "RJ", "MG", "PR", "SC", "RS", "BA", "PE", "CE", "GO", "DF"]
    canais = ["LOJA", "ONLINE", "MARKETPLACE", "TELEVENDAS"]

    n = 3500
    rows = []
    for _ in range(n):
        dt = rng.choice(datas)
        cat = rng.choice(categorias)
        marca = rng.choice(marcas[cat])
        prod = rng.choice(produtos[cat][marca])
        uf = rng.choice(ufs)
        canal = rng.choice(canais)

        qtd = int(max(1, rng.normal(4, 2)))
        preco = float(max(19.9, rng.normal(1250, 600)))
        total = qtd * preco * float(rng.uniform(0.85, 1.15))

        rows.append(
            {
                "data": dt,
                "categoria": cat,
                "marca": marca,
                "produto": prod,
                "uf": uf,
                "canal": canal,
                "quantidade": qtd,
                "preco_unitario": preco,
                "valor_total": total,
            }
        )

    df = pd.DataFrame(rows)
    return normalize_dataset(df)


def previous_period(start: date, end: date) -> Tuple[date, date]:
    days = (end - start).days + 1
    prev_end = start - timedelta(days=1)
    prev_start = prev_end - timedelta(days=days - 1)
    return prev_start, prev_end


def filter_df(
    df: pd.DataFrame,
    start: date,
    end: date,
    categoria: str,
    marca: str,
    produto: str,
    canal: str,
    uf: str,
) -> pd.DataFrame:
    dff = df.copy()

    mask = (dff["dia"] >= start) & (dff["dia"] <= end)
    dff = dff.loc[mask]

    if categoria != "Todos":
        dff = dff.loc[dff["categoria"] == categoria]
    if marca != "Todos":
        dff = dff.loc[dff["marca"] == marca]
    if produto != "Todos":
        dff = dff.loc[dff["produto"] == produto]
    if canal != "Todos":
        dff = dff.loc[dff["canal"] == canal]
    if uf != "Todos":
        dff = dff.loc[dff["uf"] == uf]

    return dff


def kpis(df_current: pd.DataFrame, df_prev: pd.DataFrame) -> Dict[str, float]:
    fat = float(df_current["valor_total"].sum())
    fat_prev = float(df_prev["valor_total"].sum())

    linhas = float(len(df_current))
    linhas_prev = float(len(df_prev))

    qtd = float(df_current["quantidade"].sum())
    qtd_prev = float(df_prev["quantidade"].sum())

    # “Pedidos”: se existir uma coluna tipo pedido/id_nf, usa; se não, usa linhas como proxy
    pedido_col = None
    for c in ["pedido", "id_pedido", "nf", "nfe", "id_nf", "numero_pedido"]:
        if c in df_current.columns:
            pedido_col = c
            break

    if pedido_col:
        pedidos = float(df_current[pedido_col].nunique())
        pedidos_prev = float(df_prev[pedido_col].nunique())
    else:
        pedidos = linhas
        pedidos_prev = linhas_prev

    ticket = (fat / pedidos) if pedidos > 0 else 0.0
    ticket_prev = (fat_prev / pedidos_prev) if pedidos_prev > 0 else 0.0

    out = {
        "faturamento": fat,
        "faturamento_prev": fat_prev,
        "pedidos": pedidos,
        "pedidos_prev": pedidos_prev,
        "quantidade": qtd,
        "quantidade_prev": qtd_prev,
        "ticket": ticket,
        "ticket_prev": ticket_prev,
    }
    return out


def pct_delta(curr: float, prev: float) -> float:
    if prev == 0:
        return 0.0 if curr == 0 else 1.0
    return (curr - prev) / prev


def brl(x: float) -> str:
    # formata BRL (pt-BR)
    s = f"{x:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


# =========================
# EXPORTAÇÃO EXCEL (OPENPYXL)
# =========================
def _format_sheet(ws, currency_cols: List[int], date_cols: List[int], percent_cols: List[int]) -> None:
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1B2A41")
    header_font = Font(bold=True, color="FFFFFF")
    header_align = Alignment(horizontal="center", vertical="center")

    # header
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # formatos
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.value is None:
                continue
            c = cell.column

            if c in currency_cols:
                cell.number_format = 'R$ #,##0.00'
            if c in percent_cols:
                cell.number_format = '0.00%'
            if c in date_cols:
                cell.number_format = 'yyyy-mm-dd'

    # largura automática
    for col in ws.columns:
        max_len = 10
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                v = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(v))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)


def build_excel_bytes(
    df_filtered: pd.DataFrame,
    df_daily: pd.DataFrame,
    df_topn: pd.DataFrame,
    kpi_table: pd.DataFrame,
) -> bytes:
    from openpyxl import load_workbook

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        kpi_table.to_excel(writer, index=False, sheet_name="KPIs")
        df_topn.to_excel(writer, index=False, sheet_name="TopN")
        df_daily.to_excel(writer, index=False, sheet_name="Serie_Diaria")
        df_filtered.to_excel(writer, index=False, sheet_name="Dados_Filtrados")

    output.seek(0)
    wb = load_workbook(output)

    # KPIs
    ws = wb["KPIs"]
    _format_sheet(ws, currency_cols=[2, 4], date_cols=[], percent_cols=[3, 5])

    # TopN
    ws = wb["TopN"]
    # 2ª coluna normalmente valor
    _format_sheet(ws, currency_cols=[2], date_cols=[], percent_cols=[])

    # Série diária
    ws = wb["Serie_Diaria"]
    # 1ª coluna data, 2ª valor
    _format_sheet(ws, currency_cols=[2, 3], date_cols=[1], percent_cols=[])

    # Dados filtrados
    ws = wb["Dados_Filtrados"]
    # tenta achar colunas relevantes
    header = [c.value for c in ws[1]]
    currency_cols = []
    date_cols = []
    for i, name in enumerate(header, start=1):
        if name in ["valor_total", "preco_unitario"]:
            currency_cols.append(i)
        if name in ["data", "dia"]:
            date_cols.append(i)

    _format_sheet(ws, currency_cols=currency_cols, date_cols=date_cols, percent_cols=[])

    out2 = BytesIO()
    wb.save(out2)
    return out2.getvalue()


# =========================
# SESSION STATE
# =========================
def init_state(df: pd.DataFrame) -> None:
    min_d = df["dia"].min()
    max_d = df["dia"].max()

    defaults = {
        "use_demo": True,
        "date_start": min_d,
        "date_end": max_d,
        "categoria": "Todos",
        "marca": "Todos",
        "produto": "Todos",
        "canal": "Todos",
        "uf": "Todos",
        "topn": 10,
        "drill_mode": "Sunburst (drill-down)",
        "group_level": "Categoria",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


def reset_filters(df: pd.DataFrame) -> None:
    min_d = df["dia"].min()
    max_d = df["dia"].max()

    st.session_state["date_start"] = min_d
    st.session_state["date_end"] = max_d
    st.session_state["categoria"] = "Todos"
    st.session_state["marca"] = "Todos"
    st.session_state["produto"] = "Todos"
    st.session_state["canal"] = "Todos"
    st.session_state["uf"] = "Todos"
    st.session_state["topn"] = 10


# =========================
# UI - HEADER
# =========================
st.markdown('<div class="pbi-title">📊 Dashboard Executivo (Power BI Premium v3)</div>', unsafe_allow_html=True)
st.markdown(
    '<div class="pbi-sub">Tema escuro • Menu recolhível • Drill-down hierárquico • Top N dinâmico • Período anterior • Exportação Excel multi-abas</div>',
    unsafe_allow_html=True,
)


# =========================
# DADOS (UPLOAD / DEMO)
# =========================
with st.sidebar:
    st.markdown("## ⚙️ Configurações")

    col_reset, col_sp = st.columns([1, 0.01])
    with col_reset:
        if st.button("🔄 Reset filtros (Power BI)", use_container_width=True):
            # usa demo provisoriamente (será substituído logo abaixo)
            df_tmp = demo_data()
            reset_filters(df_tmp)
            st.rerun()

    st.divider()
    use_demo = st.toggle("Usar base DEMO (recomendado)", value=st.session_state.get("use_demo", True))
    st.session_state["use_demo"] = use_demo

    uploaded = st.file_uploader("📤 Enviar CSV/Excel", type=["csv", "xlsx", "xls"])

# carrega df
if st.session_state["use_demo"] or uploaded is None:
    df = demo_data()
else:
    try:
        if uploaded.name.lower().endswith(".csv"):
            df_raw = pd.read_csv(uploaded, sep=None, engine="python")
        else:
            df_raw = pd.read_excel(uploaded)
        df = normalize_dataset(df_raw)
    except Exception as e:
        st.error(f"Falha ao ler o arquivo: {e}")
        st.stop()

init_state(df)


# =========================
# FILTROS (HIERARQUIA)
# =========================
min_d = df["dia"].min()
max_d = df["dia"].max()

cats = ["Todos"] + sorted(df["categoria"].unique().tolist())

# categoria -> marca -> produto
if st.session_state["categoria"] == "Todos":
    marcas_base = df
else:
    marcas_base = df[df["categoria"] == st.session_state["categoria"]]
brands = ["Todos"] + sorted(marcas_base["marca"].unique().tolist())

if st.session_state["marca"] == "Todos":
    prod_base = marcas_base
else:
    prod_base = marcas_base[marcas_base["marca"] == st.session_state["marca"]]
prods = ["Todos"] + sorted(prod_base["produto"].unique().tolist())

channels = ["Todos"] + sorted(df["canal"].unique().tolist())
ufs = ["Todos"] + sorted(df["uf"].unique().tolist())

with st.sidebar:
    st.markdown("## 🎚️ Filtros (Slicers)")

    st.session_state["date_start"] = st.date_input(
        "📅 Data inicial",
        value=st.session_state["date_start"],
        min_value=min_d,
        max_value=max_d,
    )
    st.session_state["date_end"] = st.date_input(
        "📅 Data final",
        value=st.session_state["date_end"],
        min_value=min_d,
        max_value=max_d,
    )

    st.session_state["categoria"] = st.selectbox("Categoria", cats, index=cats.index(st.session_state["categoria"]))
    st.session_state["marca"] = st.selectbox("Marca", brands, index=brands.index(st.session_state["marca"]))
    st.session_state["produto"] = st.selectbox("Produto", prods, index=prods.index(st.session_state["produto"]))
    st.session_state["canal"] = st.selectbox("Canal", channels, index=channels.index(st.session_state["canal"]))
    st.session_state["uf"] = st.selectbox("UF", ufs, index=ufs.index(st.session_state["uf"]))

    st.divider()
    st.session_state["topn"] = st.slider("Top N (dinâmico)", 5, 30, int(st.session_state["topn"]))
    st.session_state["drill_mode"] = st.radio("Drill-down", ["Sunburst (drill-down)", "Treemap (drill-down)"], index=0)


# valida datas
start = st.session_state["date_start"]
end = st.session_state["date_end"]
if start > end:
    st.error("Data inicial não pode ser maior que a data final.")
    st.stop()

prev_s, prev_e = previous_period(start, end)


df_cur = filter_df(
    df,
    start,
    end,
    st.session_state["categoria"],
    st.session_state["marca"],
    st.session_state["produto"],
    st.session_state["canal"],
    st.session_state["uf"],
)
df_prev = filter_df(
    df,
    prev_s,
    prev_e,
    st.session_state["categoria"],
    st.session_state["marca"],
    st.session_state["produto"],
    st.session_state["canal"],
    st.session_state["uf"],
)

k = kpis(df_cur, df_prev)

# =========================
# KPI CARDS
# =========================
kpi_cols = st.columns(4)
with kpi_cols[0]:
    st.metric(
        "Faturamento",
        brl(k["faturamento"]),
        delta=f"{pct_delta(k['faturamento'], k['faturamento_prev'])*100:.1f}%",
    )
with kpi_cols[1]:
    st.metric(
        "Pedidos (proxy)",
        f"{int(k['pedidos']):,}".replace(",", "."),
        delta=f"{pct_delta(k['pedidos'], k['pedidos_prev'])*100:.1f}%",
    )
with kpi_cols[2]:
    st.metric(
        "Quantidade",
        f"{int(k['quantidade']):,}".replace(",", "."),
        delta=f"{pct_delta(k['quantidade'], k['quantidade_prev'])*100:.1f}%",
    )
with kpi_cols[3]:
    st.metric(
        "Ticket Médio",
        brl(k["ticket"]),
        delta=f"{pct_delta(k['ticket'], k['ticket_prev'])*100:.1f}%",
    )

st.caption(
    f"Comparativo automático do período anterior: **{prev_s} → {prev_e}** (mesmo número de dias)."
)

# =========================
# TABS
# =========================
tab1, tab2, tab3, tab4, tab5 = st.tabs(
    ["🏠 Visão Geral", "📉 Drill-down (Hierarquia)", "🏆 Ranking Top N", "🧾 Comparativo", "📦 Exportação Excel"]
)

# =========
# VISÃO GERAL
# =========
with tab1:
    left, right = st.columns([1.6, 1.0])

    # Série diária
    df_daily = (
        df_cur.groupby("dia", as_index=False)["valor_total"].sum()
        .sort_values("dia")
    )
    df_daily["media_movel_7d"] = df_daily["valor_total"].rolling(7, min_periods=1).mean()

    fig_line = px.line(
        df_daily,
        x="dia",
        y=["valor_total", "media_movel_7d"],
        title="Evolução do Faturamento (diário)",
        labels={"value": "Valor", "variable": "Métrica", "dia": "Data"},
    )
    fig_line.update_layout(height=420, template="plotly_dark", legend_title_text="")
    fig_line.update_traces(mode="lines")

    # Ranking categorias (Top 12 fixo)
    df_cat = (
        df_cur.groupby("categoria", as_index=False)["valor_total"].sum()
        .sort_values("valor_total", ascending=False)
        .head(12)
    )
    fig_cat = px.bar(
        df_cat.sort_values("valor_total"),
        x="valor_total",
        y="categoria",
        orientation="h",
        title="Valor por Categoria (Top 12)",
        labels={"valor_total": "Faturamento", "categoria": "Categoria"},
    )
    fig_cat.update_layout(height=420, template="plotly_dark")

    with left:
        st.plotly_chart(fig_line, use_container_width=True)
    with right:
        st.plotly_chart(fig_cat, use_container_width=True)

    st.markdown("### 📋 Amostra do dataset filtrado")
    st.dataframe(df_cur.head(50), use_container_width=True)


# =========
# DRILL-DOWN (Categoria → Marca → Produto)
# =========
with tab2:
    st.markdown("### 🧩 Drill-down por hierarquia (clicável)")

    df_h = df_cur.copy()
    df_h["valor_total"] = df_h["valor_total"].astype(float)

    if len(df_h) == 0:
        st.warning("Sem dados para os filtros selecionados.")
    else:
        if st.session_state["drill_mode"].startswith("Sunburst"):
            fig_drill = px.sunburst(
                df_h,
                path=["categoria", "marca", "produto"],
                values="valor_total",
                title="Hierarquia de Faturamento (Categoria → Marca → Produto)",
            )
        else:
            fig_drill = px.treemap(
                df_h,
                path=["categoria", "marca", "produto"],
                values="valor_total",
                title="Hierarquia de Faturamento (Treemap)",
            )

        fig_drill.update_layout(template="plotly_dark", height=640)
        st.plotly_chart(fig_drill, use_container_width=True)

        st.info("Dica: clique nos blocos do gráfico para “descer nível” (drill-down), igual Power BI.")


# =========
# TOP N DINÂMICO
# =========
with tab3:
    st.markdown("### 🏆 Ranking Top N (dinâmico igual Power BI)")

    level = st.selectbox(
        "Ranking por nível",
        ["Categoria", "Marca", "Produto", "UF", "Canal"],
        index=2,
    )
    level_map = {
        "Categoria": "categoria",
        "Marca": "marca",
        "Produto": "produto",
        "UF": "uf",
        "Canal": "canal",
    }
    col_level = level_map[level]

    topn = int(st.session_state["topn"])

    df_topn = (
        df_cur.groupby(col_level, as_index=False)["valor_total"].sum()
        .sort_values("valor_total", ascending=False)
        .head(topn)
    )

    fig_top = px.bar(
        df_topn.sort_values("valor_total"),
        x="valor_total",
        y=col_level,
        orientation="h",
        title=f"Top {topn} - {level} (por faturamento)",
        labels={"valor_total": "Faturamento", col_level: level},
    )
    fig_top.update_layout(template="plotly_dark", height=560)
    st.plotly_chart(fig_top, use_container_width=True)

    st.dataframe(df_topn, use_container_width=True)


# =========
# COMPARATIVO - PERÍODO ANTERIOR
# =========
with tab4:
    st.markdown("### 🧾 Comparativo: período atual vs período anterior")

    # série atual vs anterior (diário)
    cur_daily = df_cur.groupby("dia", as_index=False)["valor_total"].sum().sort_values("dia")
    prev_daily = df_prev.groupby("dia", as_index=False)["valor_total"].sum().sort_values("dia")

    cur_daily = cur_daily.rename(columns={"valor_total": "atual"})
    prev_daily = prev_daily.rename(columns={"valor_total": "anterior"})

    # cria eixo contínuo do período atual
    full_days = pd.date_range(start, end, freq="D").date
    base = pd.DataFrame({"dia": full_days})

    cur_daily = base.merge(cur_daily, on="dia", how="left").fillna({"atual": 0})
    # para comparar com o anterior, alinhe pelo "dia relativo"
    prev_days = pd.date_range(prev_s, prev_e, freq="D").date
    prev_base = pd.DataFrame({"dia_prev": prev_days})
    prev_base["idx"] = range(len(prev_base))

    cur_daily["idx"] = range(len(cur_daily))
    prev_daily = prev_base.merge(prev_daily, left_on="dia_prev", right_on="dia", how="left").fillna({"anterior": 0})
    prev_daily = prev_daily[["idx", "anterior"]]

    comp = cur_daily.merge(prev_daily, on="idx", how="left").fillna({"anterior": 0})

    fig_comp = px.line(
        comp,
        x="dia",
        y=["atual", "anterior"],
        title="Faturamento diário: Atual vs Período anterior (mesma duração)",
        labels={"value": "Faturamento", "variable": "Série", "dia": "Data"},
    )
    fig_comp.update_layout(template="plotly_dark", height=520)
    st.plotly_chart(fig_comp, use_container_width=True)

    st.markdown("#### KPIs detalhados (Atual x Anterior)")
    kpi_table = pd.DataFrame(
        [
            ["Faturamento", k["faturamento"], k["faturamento_prev"], pct_delta(k["faturamento"], k["faturamento_prev"])],
            ["Pedidos (proxy)", k["pedidos"], k["pedidos_prev"], pct_delta(k["pedidos"], k["pedidos_prev"])],
            ["Quantidade", k["quantidade"], k["quantidade_prev"], pct_delta(k["quantidade"], k["quantidade_prev"])],
            ["Ticket Médio", k["ticket"], k["ticket_prev"], pct_delta(k["ticket"], k["ticket_prev"])],
        ],
        columns=["Indicador", "Atual", "Anterior", "Variação_%"],
    )
    st.dataframe(kpi_table, use_container_width=True)


# =========
# EXPORTAÇÃO EXCEL MULTI-ABAS + FORMATAÇÃO
# =========
with tab5:
    st.markdown("### 📦 Exportação Excel (Multi-abas + formatação)")

    if len(df_cur) == 0:
        st.warning("Sem dados para exportar com os filtros atuais.")
    else:
        # prepara tabelas para export
        df_daily_export = df_daily.copy()
        df_daily_export["dia"] = pd.to_datetime(df_daily_export["dia"])
        df_daily_export = df_daily_export.rename(columns={"valor_total": "total_diario"})

        # TopN atual do nível selecionado do tab3? aqui vai top por produto como padrão
        df_topn_export = (
            df_cur.groupby("produto", as_index=False)["valor_total"].sum()
            .sort_values("valor_total", ascending=False)
            .head(int(st.session_state["topn"]))
        )

        kpi_table_export = pd.DataFrame(
            [
                ["Faturamento", k["faturamento"], k["faturamento_prev"], pct_delta(k["faturamento"], k["faturamento_prev"])],
                ["Pedidos (proxy)", k["pedidos"], k["pedidos_prev"], pct_delta(k["pedidos"], k["pedidos_prev"])],
                ["Quantidade", k["quantidade"], k["quantidade_prev"], pct_delta(k["quantidade"], k["quantidade_prev"])],
                ["Ticket Médio", k["ticket"], k["ticket_prev"], pct_delta(k["ticket"], k["ticket_prev"])],
            ],
            columns=["Indicador", "Atual", "Anterior", "Variação_%"],
        )

        # bytes excel
        excel_bytes = build_excel_bytes(
            df_filtered=df_cur,
            df_daily=df_daily_export[["dia", "total_diario", "media_movel_7d"]],
            df_topn=df_topn_export,
            kpi_table=kpi_table_export,
        )

        ts_name = datetime.now().strftime("%Y%m%d_%H%M%S")
        file_name = f"dashboard_pbi_premium_v3_SAFE_{ts_name}.xlsx"

        st.download_button(
            "⬇️ Baixar Excel (com abas + formatação)",
            data=excel_bytes,
            file_name=file_name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        st.info(
            "Exportação SAFE usando **openpyxl** (não precisa instalar xlsxwriter). "
            "Abas: KPIs • TopN • Série_Diária • Dados_Filtrados."
        )