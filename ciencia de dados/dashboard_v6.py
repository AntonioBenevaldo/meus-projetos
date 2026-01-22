import sys
from pathlib import Path

import pandas as pd

from PySide6.QtCore import Qt, QAbstractTableModel, QDate
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFileDialog,
    QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QLineEdit, QTabWidget, QTableView, QMessageBox,
    QComboBox, QGroupBox, QFormLayout, QDateEdit,
    QFrame, QGridLayout, QSizePolicy, QTextEdit, QSpacerItem
)

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference


# =========================================================
# Utils
# =========================================================
def format_brl(valor: float) -> str:
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    base = df.copy()

    # Produto
    if "produto" not in base.columns and "descricao" in base.columns:
        base.rename(columns={"descricao": "produto"}, inplace=True)

    # Canal
    if "canal" not in base.columns and "canal_aquisicao" in base.columns:
        base.rename(columns={"canal_aquisicao": "canal"}, inplace=True)

    # Data
    if "data" not in base.columns and "dh_emissao" in base.columns:
        base.rename(columns={"dh_emissao": "data"}, inplace=True)

    if "data" in base.columns:
        base["data"] = pd.to_datetime(base["data"], errors="coerce")

    # Valores
    if "valor_total" not in base.columns and "v_prod" in base.columns:
        base.rename(columns={"v_prod": "valor_total"}, inplace=True)

    if "preco_unitario" not in base.columns and "v_un" in base.columns:
        base.rename(columns={"v_un": "preco_unitario"}, inplace=True)

    # EAN
    if "ean" in base.columns:
        base["ean"] = base["ean"].astype(str)

    # Nome cliente
    if "nome" not in base.columns:
        for cand in ["razao_social", "cliente", "destinatario_nome"]:
            if cand in base.columns:
                base.rename(columns={cand: "nome"}, inplace=True)
                break

    return base


def checar_arquivos_erp(data_dir: Path):
    arquivos = ["clientes.csv", "produtos.csv", "vendas.csv", "itens_venda.csv"]
    return [a for a in arquivos if not (data_dir / a).exists()]


def carregar_base_erp(data_dir: Path) -> pd.DataFrame:
    clientes = pd.read_csv(data_dir / "clientes.csv")
    produtos = pd.read_csv(data_dir / "produtos.csv")
    vendas = pd.read_csv(data_dir / "vendas.csv", parse_dates=["data"])
    itens = pd.read_csv(data_dir / "itens_venda.csv")

    base = (itens
            .merge(vendas, on="venda_id", how="left")
            .merge(produtos, on="produto_id", how="left")
            .merge(clientes, on="cliente_id", how="left"))

    return normalize_cols(base)


def compute_outlier_mask(series: pd.Series) -> pd.Series:
    if series is None or series.empty:
        return pd.Series([False] * 0)

    s = pd.to_numeric(series, errors="coerce").dropna()
    if s.empty:
        return pd.Series([False] * len(series), index=series.index)

    q1 = s.quantile(0.25)
    q3 = s.quantile(0.75)
    iqr = q3 - q1
    lim_sup = q3 + 1.5 * iqr
    return pd.to_numeric(series, errors="coerce") > lim_sup


def add_excel_format_and_charts(path_xlsx: str):
    import openpyxl

    wb = openpyxl.load_workbook(path_xlsx)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        max_col = ws.max_column
        max_row = ws.max_row

        # Cabeçalho
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # Ajuste largura
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            max_len = 10
            for row in range(1, min(max_row, 200) + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                max_len = max(max_len, min(len(str(v)), 50))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 42)

        # Formato moeda
        for col in range(1, max_col + 1):
            header = str(ws.cell(row=1, column=col).value).lower()
            if any(k in header for k in ["valor", "fatur", "custo", "preco", "total", "lucro", "impacto"]):
                for row in range(2, max_row + 1):
                    c = ws.cell(row=row, column=col)
                    if isinstance(c.value, (int, float)):
                        c.number_format = '"R$" #,##0.00'
            if any(k in header for k in ["margem", "pct", "percent"]):
                for row in range(2, max_row + 1):
                    c = ws.cell(row=row, column=col)
                    if isinstance(c.value, (int, float)):
                        c.number_format = '0.00%'

    # Gráficos (se existirem abas padrão)
    def chart_line(sheet_name, title):
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        if ws.max_row < 2 or ws.max_column < 2:
            return
        chart = LineChart()
        chart.title = title
        chart.y_axis.title = "R$"
        chart.x_axis.title = "Período"
        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 11
        chart.width = 26
        ws.add_chart(chart, "D2")

    def chart_bar(sheet_name, title, x_title):
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        if ws.max_row < 2 or ws.max_column < 2:
            return
        chart = BarChart()
        chart.title = title
        chart.y_axis.title = "R$"
        chart.x_axis.title = x_title
        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 11
        chart.width = 26
        ws.add_chart(chart, "D2")

    chart_line("Faturamento_Mensal", "Faturamento Mensal")
    chart_bar("Top_Produtos", "Top Produtos", "Produto")
    chart_bar("Faturamento_Canal", "Faturamento por Canal", "Canal")

    wb.save(path_xlsx)


# =========================================================
# Table Model (com destaque)
# =========================================================
class PandasTableModel(QAbstractTableModel):
    def __init__(self, df: pd.DataFrame):
        super().__init__()
        self.df = df.copy()
        self.flag_outlier = pd.Series([False] * len(df))
        self.flag_ean_invalid = pd.Series([False] * len(df))
        self.flag_cancel = pd.Series([False] * len(df))

    def rowCount(self, parent=None):
        return len(self.df)

    def columnCount(self, parent=None):
        return len(self.df.columns)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None

        r = index.row()
        c = index.column()

        if role == Qt.DisplayRole:
            value = self.df.iloc[r, c]
            if pd.isna(value):
                return ""
            if isinstance(value, float):
                return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            return str(value)

        # cores na linha
        if role == Qt.BackgroundRole:
            try:
                if len(self.flag_cancel) > 0 and bool(self.flag_cancel.iloc[r]):
                    return QColor(255, 220, 220)  # cancel/deneg
                if len(self.flag_outlier) > 0 and bool(self.flag_outlier.iloc[r]):
                    return QColor(255, 245, 204)  # outlier
                if len(self.flag_ean_invalid) > 0 and bool(self.flag_ean_invalid.iloc[r]):
                    return QColor(220, 235, 255)  # ean inválido
            except Exception:
                return None

        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return str(self.df.columns[section])
        return str(section)

    def update(self, df: pd.DataFrame, flag_outlier=None, flag_ean_invalid=None, flag_cancel=None):
        self.beginResetModel()
        self.df = df.copy()

        n = len(df)
        self.flag_outlier = flag_outlier if flag_outlier is not None else pd.Series([False] * n)
        self.flag_ean_invalid = flag_ean_invalid if flag_ean_invalid is not None else pd.Series([False] * n)
        self.flag_cancel = flag_cancel if flag_cancel is not None else pd.Series([False] * n)

        self.flag_outlier = self.flag_outlier.reindex(df.index, fill_value=False)
        self.flag_ean_invalid = self.flag_ean_invalid.reindex(df.index, fill_value=False)
        self.flag_cancel = self.flag_cancel.reindex(df.index, fill_value=False)

        self.endResetModel()


# =========================================================
# KPI Card (Power BI style)
# =========================================================
class KpiCard(QFrame):
    def __init__(self, titulo: str, valor: str = "-", subtitulo: str = ""):
        super().__init__()
        self.setObjectName("KpiCard")
        self.setMinimumHeight(95)
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        layout = QVBoxLayout(self)
        layout.setContentsMargins(14, 12, 14, 12)
        layout.setSpacing(6)

        self.lbl_titulo = QLabel(titulo)
        self.lbl_titulo.setStyleSheet("font-size: 12px; color: #5A5A5A; font-weight: 600;")

        self.lbl_valor = QLabel(valor)
        self.lbl_valor.setStyleSheet("font-size: 22px; font-weight: 800; color: #111111;")

        self.lbl_sub = QLabel(subtitulo)
        self.lbl_sub.setStyleSheet("font-size: 11px; color: #6B7280;")

        layout.addWidget(self.lbl_titulo)
        layout.addWidget(self.lbl_valor)
        layout.addWidget(self.lbl_sub)

        self.setStyleSheet("""
            QFrame#KpiCard {
                background: #FFFFFF;
                border: 1px solid #E6E6E6;
                border-radius: 14px;
            }
        """)

    def set_value(self, valor: str, subtitulo: str = ""):
        self.lbl_valor.setText(valor)
        if subtitulo is not None:
            self.lbl_sub.setText(subtitulo)


# =========================================================
# Dashboard v6
# =========================================================
class DashboardV6(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Dashboard v6 - Resumo Executivo (Power BI Style) | Portfólio Premium")
        self.resize(1480, 850)

        self.base_dir = Path(__file__).resolve().parent
        self.data_dir = self.base_dir / "dados"

        self.base = pd.DataFrame()
        self.base_filtrada = pd.DataFrame()

        # -------------------------
        # Layout principal
        # -------------------------
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)

        # Barra superior
        top_bar = QHBoxLayout()
        self.btn_carregar = QPushButton("Carregar (pasta dados/)")
        self.btn_carregar.clicked.connect(self.on_carregar)

        self.btn_escolher = QPushButton("Escolher pasta...")
        self.btn_escolher.clicked.connect(self.on_escolher_pasta)

        self.btn_limpar = QPushButton("Limpar filtros")
        self.btn_limpar.clicked.connect(self.on_limpar)
        self.btn_limpar.setEnabled(False)

        self.btn_exportar = QPushButton("Exportar Excel v6")
        self.btn_exportar.clicked.connect(self.on_exportar)
        self.btn_exportar.setEnabled(False)

        self.lbl_status = QLabel("Status: aguardando...")
        self.lbl_status.setStyleSheet("font-weight:bold;")

        top_bar.addWidget(self.btn_carregar)
        top_bar.addWidget(self.btn_escolher)
        top_bar.addWidget(self.btn_limpar)
        top_bar.addWidget(self.btn_exportar)
        top_bar.addStretch()
        top_bar.addWidget(self.lbl_status)
        main_layout.addLayout(top_bar)

        # -------------------------
        # Filtros
        # -------------------------
        filtros_box = QGroupBox("Filtros")
        filtros_layout = QHBoxLayout(filtros_box)

        self.input_busca = QLineEdit()
        self.input_busca.setPlaceholderText("Buscar (produto, EAN, UF, canal, cliente, CFOP, CST...)")
        self.input_busca.textChanged.connect(self.aplicar_filtros)

        self.combo_uf = QComboBox()
        self.combo_uf.currentIndexChanged.connect(self.aplicar_filtros)

        self.combo_canal = QComboBox()
        self.combo_canal.currentIndexChanged.connect(self.aplicar_filtros)

        self.combo_categoria = QComboBox()
        self.combo_categoria.currentIndexChanged.connect(self.aplicar_filtros)

        self.combo_status = QComboBox()
        self.combo_status.currentIndexChanged.connect(self.aplicar_filtros)

        self.date_ini = QDateEdit()
        self.date_ini.setCalendarPopup(True)
        self.date_ini.dateChanged.connect(self.aplicar_filtros)

        self.date_fim = QDateEdit()
        self.date_fim.setCalendarPopup(True)
        self.date_fim.dateChanged.connect(self.aplicar_filtros)

        self.lbl_linhas = QLabel("Linhas: 0")
        self.lbl_linhas.setStyleSheet("font-weight:bold;")

        filtros_layout.addWidget(QLabel("Texto:"))
        filtros_layout.addWidget(self.input_busca, 2)
        filtros_layout.addWidget(self.combo_uf)
        filtros_layout.addWidget(self.combo_canal)
        filtros_layout.addWidget(self.combo_categoria)
        filtros_layout.addWidget(self.combo_status)
        filtros_layout.addWidget(QLabel("De:"))
        filtros_layout.addWidget(self.date_ini)
        filtros_layout.addWidget(QLabel("Até:"))
        filtros_layout.addWidget(self.date_fim)
        filtros_layout.addWidget(self.lbl_linhas)

        main_layout.addWidget(filtros_box)

        # -------------------------
        # Tabs
        # -------------------------
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        # TAB 0 - Resumo Executivo (Power BI)
        self.tab_exec = QWidget()
        self.tabs.addTab(self.tab_exec, "Resumo Executivo")
        self.build_exec_tab()

        # TAB 1 - Tabela
        self.tab_tabela = QWidget()
        self.tabs.addTab(self.tab_tabela, "Tabela")
        t1 = QVBoxLayout(self.tab_tabela)
        self.table_base = QTableView()
        self.table_base.setSortingEnabled(True)
        self.model_base = PandasTableModel(pd.DataFrame())
        self.table_base.setModel(self.model_base)
        t1.addWidget(self.table_base)

        # TAB 2 - Alertas
        self.tab_alertas = QWidget()
        self.tabs.addTab(self.tab_alertas, "Alertas")
        ta = QVBoxLayout(self.tab_alertas)
        self.lbl_alertas = QLabel("Carregue os dados para ver os alertas.")
        self.lbl_alertas.setTextInteractionFlags(Qt.TextSelectableByMouse)
        ta.addWidget(self.lbl_alertas)
        self.table_alertas = QTableView()
        self.table_alertas.setSortingEnabled(True)
        self.model_alertas = PandasTableModel(pd.DataFrame())
        self.table_alertas.setModel(self.model_alertas)
        ta.addWidget(self.table_alertas)

        # TAB 3 - Qualidade
        self.tab_quality = QWidget()
        self.tabs.addTab(self.tab_quality, "Qualidade de Dados")
        tq = QVBoxLayout(self.tab_quality)
        self.lbl_quality = QLabel("Carregue os dados para ver a auditoria.")
        self.lbl_quality.setTextInteractionFlags(Qt.TextSelectableByMouse)
        tq.addWidget(self.lbl_quality)

        # Autoload
        if self.data_dir.exists():
            self.on_carregar()

    # =========================================================
    # Construção do Resumo Executivo
    # =========================================================
    def build_exec_tab(self):
        layout = QVBoxLayout(self.tab_exec)
        layout.setSpacing(12)

        # Título
        titulo = QLabel("Resumo Executivo (Power BI Style)")
        titulo.setStyleSheet("font-size: 16px; font-weight: 800;")
        layout.addWidget(titulo)

        # Grid KPI Cards
        kpi_grid = QGridLayout()
        kpi_grid.setHorizontalSpacing(12)
        kpi_grid.setVerticalSpacing(12)

        self.card_fat_liq = KpiCard("Faturamento Líquido", "-")
        self.card_fat_bruto = KpiCard("Faturamento Bruto", "-")
        self.card_ticket = KpiCard("Ticket Médio", "-")
        self.card_vendas = KpiCard("Qtd Vendas", "-")
        self.card_itens = KpiCard("Qtd Itens", "-")
        self.card_margem = KpiCard("Margem Bruta", "-")
        self.card_impacto = KpiCard("Impacto Cancel/Deneg", "-")

        kpi_grid.addWidget(self.card_fat_liq, 0, 0)
        kpi_grid.addWidget(self.card_fat_bruto, 0, 1)
        kpi_grid.addWidget(self.card_ticket, 0, 2)
        kpi_grid.addWidget(self.card_vendas, 1, 0)
        kpi_grid.addWidget(self.card_itens, 1, 1)
        kpi_grid.addWidget(self.card_margem, 1, 2)
        kpi_grid.addWidget(self.card_impacto, 2, 0, 1, 3)

        layout.addLayout(kpi_grid)

        # Área de gráficos e insights
        bottom_grid = QGridLayout()
        bottom_grid.setHorizontalSpacing(12)
        bottom_grid.setVerticalSpacing(12)

        # Gráfico mensal
        self.frame_mes = self.make_chart_frame("Faturamento Mensal")
        self.fig_mes = Figure(figsize=(5, 3))
        self.canvas_mes = FigureCanvas(self.fig_mes)
        self.frame_mes["content"].addWidget(self.canvas_mes)

        # Top produtos
        self.frame_top = self.make_chart_frame("Top Produtos")
        self.fig_top = Figure(figsize=(5, 3))
        self.canvas_top = FigureCanvas(self.fig_top)
        self.frame_top["content"].addWidget(self.canvas_top)

        # Por canal
        self.frame_canal = self.make_chart_frame("Faturamento por Canal")
        self.fig_canal = Figure(figsize=(5, 3))
        self.canvas_canal = FigureCanvas(self.fig_canal)
        self.frame_canal["content"].addWidget(self.canvas_canal)

        # Insights
        self.frame_insights = self.make_chart_frame("Insights Automáticos")
        self.txt_insights = QTextEdit()
        self.txt_insights.setReadOnly(True)
        self.txt_insights.setStyleSheet("""
            QTextEdit {
                background: #FFFFFF;
                border: 1px solid #E6E6E6;
                border-radius: 12px;
                padding: 10px;
                font-size: 12px;
            }
        """)
        self.frame_insights["content"].addWidget(self.txt_insights)

        bottom_grid.addWidget(self.frame_mes["frame"], 0, 0)
        bottom_grid.addWidget(self.frame_top["frame"], 0, 1)
        bottom_grid.addWidget(self.frame_canal["frame"], 1, 0)
        bottom_grid.addWidget(self.frame_insights["frame"], 1, 1)

        layout.addLayout(bottom_grid)

    def make_chart_frame(self, title: str):
        frame = QFrame()
        frame.setStyleSheet("""
            QFrame {
                background: #FFFFFF;
                border: 1px solid #E6E6E6;
                border-radius: 14px;
            }
        """)
        outer = QVBoxLayout(frame)
        outer.setContentsMargins(12, 10, 12, 10)
        outer.setSpacing(8)

        lbl = QLabel(title)
        lbl.setStyleSheet("font-weight: 800; font-size: 12px; color: #111111;")
        outer.addWidget(lbl)

        content = QVBoxLayout()
        content.setContentsMargins(0, 0, 0, 0)
        outer.addLayout(content)

        return {"frame": frame, "content": content}

    # =========================================================
    # Carregamento
    # =========================================================
    def on_escolher_pasta(self):
        folder = QFileDialog.getExistingDirectory(self, "Selecione a pasta dos CSVs")
        if folder:
            self.data_dir = Path(folder)
            self.on_carregar()

    def on_carregar(self):
        try:
            faltando = checar_arquivos_erp(self.data_dir)
            if faltando:
                raise FileNotFoundError(f"Arquivos faltando em {self.data_dir}: {faltando}")

            self.base = carregar_base_erp(self.data_dir)
            self.base_filtrada = self.base.copy()

            self.preencher_combos()
            self.preencher_periodo()
            self.atualizar_tudo()

            self.btn_limpar.setEnabled(True)
            self.btn_exportar.setEnabled(True)

            self.lbl_status.setText(f"Status: OK | Base: {self.base.shape[0]} linhas / {self.base.shape[1]} colunas")

        except Exception as e:
            QMessageBox.critical(self, "Erro ao carregar", str(e))
            self.lbl_status.setText("Status: ERRO ao carregar")

    def preencher_combos(self):
        self.combo_uf.blockSignals(True)
        self.combo_canal.blockSignals(True)
        self.combo_categoria.blockSignals(True)
        self.combo_status.blockSignals(True)

        self.combo_uf.clear()
        self.combo_canal.clear()
        self.combo_categoria.clear()
        self.combo_status.clear()

        self.combo_uf.addItem("UF: Todas")
        self.combo_canal.addItem("Canal: Todos")
        self.combo_categoria.addItem("Categoria: Todas")
        self.combo_status.addItem("Status: Todos")

        if "uf_destino" in self.base.columns:
            for u in sorted(self.base["uf_destino"].dropna().unique().tolist()):
                self.combo_uf.addItem(f"UF: {u}")

        if "canal" in self.base.columns:
            for c in sorted(self.base["canal"].dropna().unique().tolist()):
                self.combo_canal.addItem(f"Canal: {c}")

        if "categoria" in self.base.columns:
            for cat in sorted(self.base["categoria"].dropna().unique().tolist()):
                self.combo_categoria.addItem(f"Categoria: {cat}")

        if "status" in self.base.columns:
            for st in sorted(self.base["status"].dropna().unique().tolist()):
                self.combo_status.addItem(f"Status: {st}")

        self.combo_uf.blockSignals(False)
        self.combo_canal.blockSignals(False)
        self.combo_categoria.blockSignals(False)
        self.combo_status.blockSignals(False)

    def preencher_periodo(self):
        if "data" not in self.base.columns:
            hoje = QDate.currentDate()
            self.date_ini.setDate(hoje.addMonths(-6))
            self.date_fim.setDate(hoje)
            return

        d = self.base["data"].dropna()
        if d.empty:
            hoje = QDate.currentDate()
            self.date_ini.setDate(hoje.addMonths(-6))
            self.date_fim.setDate(hoje)
            return

        min_dt = d.min().date()
        max_dt = d.max().date()

        self.date_ini.blockSignals(True)
        self.date_fim.blockSignals(True)
        self.date_ini.setDate(QDate(min_dt.year, min_dt.month, min_dt.day))
        self.date_fim.setDate(QDate(max_dt.year, max_dt.month, max_dt.day))
        self.date_ini.blockSignals(False)
        self.date_fim.blockSignals(False)

    # =========================================================
    # Filtros
    # =========================================================
    def on_limpar(self):
        self.input_busca.setText("")
        self.combo_uf.setCurrentIndex(0)
        self.combo_canal.setCurrentIndex(0)
        self.combo_categoria.setCurrentIndex(0)
        self.combo_status.setCurrentIndex(0)
        self.preencher_periodo()

        self.base_filtrada = self.base.copy()
        self.atualizar_tudo()

    def aplicar_filtros(self):
        if self.base.empty:
            return

        df = self.base.copy()

        # Período
        if "data" in df.columns:
            di = self.date_ini.date().toPython()
            dfim = self.date_fim.date().toPython()

            df["data"] = pd.to_datetime(df["data"], errors="coerce")
            df = df.dropna(subset=["data"])
            df = df[(df["data"].dt.date >= di) & (df["data"].dt.date <= dfim)]

        # Texto
        termo = self.input_busca.text().strip().lower()
        if termo:
            cols_preferidas = [c for c in df.columns if c.lower() in [
                "produto", "ean", "uf_destino", "canal", "categoria",
                "cfop", "cst_icms", "nome", "marca"
            ]]
            if not cols_preferidas:
                cols_preferidas = df.columns.tolist()

            mask = df[cols_preferidas].astype(str).apply(
                lambda col: col.str.lower().str.contains(termo, na=False)
            ).any(axis=1)
            df = df[mask]

        # UF
        uf_sel = self.combo_uf.currentText().replace("UF: ", "")
        if uf_sel != "Todas" and "uf_destino" in df.columns:
            df = df[df["uf_destino"] == uf_sel]

        # Canal
        canal_sel = self.combo_canal.currentText().replace("Canal: ", "")
        if canal_sel != "Todos" and "canal" in df.columns:
            df = df[df["canal"] == canal_sel]

        # Categoria
        cat_sel = self.combo_categoria.currentText().replace("Categoria: ", "")
        if cat_sel != "Todas" and "categoria" in df.columns:
            df = df[df["categoria"] == cat_sel]

        # Status
        st_sel = self.combo_status.currentText().replace("Status: ", "")
        if st_sel != "Todos" and "status" in df.columns:
            df = df[df["status"] == st_sel]

        self.base_filtrada = df.copy()
        self.atualizar_tudo()

    # =========================================================
    # KPIs / Agregações
    # =========================================================
    def faturamento_liquido(self, df: pd.DataFrame) -> tuple[float, float, float]:
        if "valor_total" not in df.columns:
            return 0.0, 0.0, 0.0

        fat_bruto = float(df["valor_total"].sum())

        if "status" not in df.columns:
            return fat_bruto, fat_bruto, 0.0

        status = df["status"].astype(str).str.upper()
        invalidos = status.str.contains("CANCEL") | status.str.contains("DENEG") | status.str.contains("INUTIL")

        df_ok = df[~invalidos].copy()
        fat_liq = float(df_ok["valor_total"].sum())
        impacto = fat_bruto - fat_liq
        return fat_bruto, fat_liq, impacto

    def calcular_kpis(self, df: pd.DataFrame) -> dict:
        if df.empty or "valor_total" not in df.columns:
            return {}

        fat_bruto, fat_liq, impacto = self.faturamento_liquido(df)

        qtd_vendas = int(df["venda_id"].nunique()) if "venda_id" in df.columns else int(df.shape[0])
        ticket = float(df.groupby("venda_id")["valor_total"].sum().mean()) if "venda_id" in df.columns else fat_bruto
        qtd_itens = int(df.shape[0])

        margem_pct = None
        if "custo_total" in df.columns:
            custo = float(df["custo_total"].sum())
            lucro = fat_bruto - custo
            margem_pct = float(lucro / fat_bruto) if fat_bruto != 0 else 0.0

        return {
            "fat_bruto": fat_bruto,
            "fat_liq": fat_liq,
            "impacto": impacto,
            "qtd_vendas": qtd_vendas,
            "ticket": ticket,
            "qtd_itens": qtd_itens,
            "margem_pct": margem_pct
        }

    def top_produtos(self, df: pd.DataFrame) -> pd.Series:
        if "produto" not in df.columns or "valor_total" not in df.columns:
            return pd.Series(dtype=float)
        return (df.groupby("produto")["valor_total"].sum().sort_values(ascending=False).head(10))

    def serie_mensal(self, df: pd.DataFrame) -> pd.Series:
        if "data" not in df.columns or "valor_total" not in df.columns:
            return pd.Series(dtype=float)
        tmp = df.dropna(subset=["data"]).copy()
        if tmp.empty:
            return pd.Series(dtype=float)
        return (tmp.groupby(tmp["data"].dt.to_period("M"))["valor_total"].sum().sort_index())

    def por_canal(self, df: pd.DataFrame) -> pd.Series:
        if "canal" not in df.columns or "valor_total" not in df.columns:
            return pd.Series(dtype=float)
        return (df.groupby("canal")["valor_total"].sum().sort_values(ascending=False))

    # =========================================================
    # Alertas / Flags
    # =========================================================
    def gerar_flags_tabela(self, df: pd.DataFrame):
        n = len(df)
        if n == 0:
            return pd.Series([], dtype=bool), pd.Series([], dtype=bool), pd.Series([], dtype=bool)

        flag_out = pd.Series([False] * n, index=df.index)
        if "preco_unitario" in df.columns:
            flag_out = compute_outlier_mask(df["preco_unitario"]).fillna(False)

        flag_ean = pd.Series([False] * n, index=df.index)
        if "ean" in df.columns:
            flag_ean = df["ean"].astype(str).str.len().ne(13)

        flag_cancel = pd.Series([False] * n, index=df.index)
        if "status" in df.columns:
            st = df["status"].astype(str).str.upper()
            flag_cancel = st.str.contains("CANCEL") | st.str.contains("DENEG") | st.str.contains("INUTIL")

        return flag_out, flag_ean, flag_cancel

    def gerar_alertas(self, df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame()

        alerts = []

        if "status" in df.columns:
            st = df["status"].astype(str).str.upper()
            mask_cancel = st.str.contains("CANCEL") | st.str.contains("DENEG") | st.str.contains("INUTIL")
            if mask_cancel.any():
                a = df[mask_cancel].copy()
                a["tipo_alerta"] = "STATUS_CANCEL_DENEG"
                alerts.append(a)

        if "ean" in df.columns:
            ean_len = df["ean"].astype(str).str.len()
            mask_ean = ean_len != 13
            if mask_ean.any():
                a = df[mask_ean].copy()
                a["tipo_alerta"] = "EAN_INVALIDO_LEN"
                alerts.append(a)

            dup = df[df.duplicated(subset=["ean"], keep=False)].copy()
            if not dup.empty:
                dup["tipo_alerta"] = "EAN_DUPLICADO"
                alerts.append(dup)

        if "preco_unitario" in df.columns:
            mask_out = compute_outlier_mask(df["preco_unitario"])
            if mask_out.any():
                a = df[mask_out].copy()
                a["tipo_alerta"] = "OUTLIER_PRECO_UNITARIO"
                alerts.append(a)

        if not alerts:
            return pd.DataFrame()

        final = pd.concat(alerts, ignore_index=True)

        cols_show = []
        for c in ["tipo_alerta", "data", "venda_id", "produto_id", "produto", "ean",
                  "quantidade", "preco_unitario", "valor_total", "uf_destino", "cfop", "status", "nome"]:
            if c in final.columns:
                cols_show.append(c)

        if cols_show:
            final = final[cols_show]

        return final.head(1000)

    # =========================================================
    # Qualidade / Insights
    # =========================================================
    def quality_text(self, df: pd.DataFrame) -> str:
        if df.empty:
            return "🧪 Qualidade de Dados\n\nSem dados no recorte atual."

        linhas = []
        linhas.append(f"- Linhas: {df.shape[0]}")
        linhas.append(f"- Colunas: {df.shape[1]}")

        nulos = df.isna().sum().sort_values(ascending=False).head(12)
        linhas.append("\nTop colunas com nulos:")
        for col, qtd in nulos.items():
            if qtd > 0:
                linhas.append(f"  • {col}: {int(qtd)}")

        if "ean" in df.columns:
            inv = int(df["ean"].astype(str).str.len().ne(13).sum())
            dup = int(df[df.duplicated(subset=["ean"], keep=False)]["ean"].nunique())
            linhas.append(f"\nEAN inválidos (len != 13): {inv}")
            linhas.append(f"EAN duplicados (distintos): {dup}")

        if "preco_unitario" in df.columns:
            out = int(compute_outlier_mask(df["preco_unitario"]).sum())
            linhas.append(f"\nOutliers preco_unitario (IQR): {out}")

        return "🧪 Qualidade de Dados (recorte atual)\n\n" + "\n".join(linhas)

    def insights_text(self, df: pd.DataFrame) -> str:
        if df.empty or "valor_total" not in df.columns:
            return "Sem dados suficientes para gerar insights."

        k = self.calcular_kpis(df)
        top = self.top_produtos(df)
        canal = self.por_canal(df)
        alerts = self.gerar_alertas(df)

        linhas = []
        linhas.append("Resumo interpretável do recorte atual:")
        linhas.append("")

        # faturamento
        linhas.append(f"- Faturamento líquido: {format_brl(k.get('fat_liq', 0))}")
        linhas.append(f"- Faturamento bruto: {format_brl(k.get('fat_bruto', 0))}")
        if k.get("impacto", 0) > 0:
            linhas.append(f"- Impacto por cancel/deneg: {format_brl(k.get('impacto', 0))} (reduz faturamento líquido)")
        else:
            linhas.append("- Sem impacto relevante por cancel/deneg no recorte.")

        # ticket e vendas
        linhas.append(f"- Qtd vendas: {k.get('qtd_vendas', 0)} | Ticket médio: {format_brl(k.get('ticket', 0))}")

        # margem
        if k.get("margem_pct") is not None:
            mp = k["margem_pct"] * 100
            linhas.append(f"- Margem bruta estimada: {mp:.2f}%".replace(".", ","))
            if mp < 10:
                linhas.append("  • Alerta: margem baixa. Verifique desconto alto ou custo elevado por categoria/produto.")
            elif mp > 30:
                linhas.append("  • Margem alta. Ótimo, mas confirme se custos estão consistentes.")

        # concentração
        if len(top) > 0:
            total = float(df["valor_total"].sum())
            share_top1 = float(top.iloc[0] / total) if total != 0 else 0
            linhas.append("")
            linhas.append(f"- Produto líder: {top.index[0]} | participação no faturamento: {share_top1*100:.2f}%".replace(".", ","))
            if share_top1 > 0.25:
                linhas.append("  • Concentração alta em 1 produto. Avalie risco de dependência do mix.")

        # canal dominante
        if len(canal) > 0:
            canal_top = canal.index[0]
            share_canal = float(canal.iloc[0] / canal.sum()) if canal.sum() != 0 else 0
            linhas.append(f"- Canal dominante: {canal_top} ({share_canal*100:.2f}%)".replace(".", ","))
            if share_canal > 0.6:
                linhas.append("  • Dependência alta de um canal. Diversificar pode reduzir risco comercial.")

        # alertas
        if not alerts.empty:
            vc = alerts["tipo_alerta"].value_counts()
            linhas.append("")
            linhas.append("Alertas detectados:")
            for t, v in vc.items():
                linhas.append(f"- {t}: {int(v)} ocorrências")
            linhas.append("Sugestão: corrija EAN inválido e valide outliers antes de treinar modelos.")
        else:
            linhas.append("")
            linhas.append("Nenhum alerta relevante identificado no recorte atual.")

        return "\n".join(linhas)

    # =========================================================
    # Atualizações UI (Resumo + Tabela + Alertas)
    # =========================================================
    def atualizar_tudo(self):
        df = self.base_filtrada if not self.base_filtrada.empty else self.base

        # flags para tabela
        flag_out, flag_ean, flag_cancel = self.gerar_flags_tabela(df)
        self.model_base.update(df, flag_outlier=flag_out, flag_ean_invalid=flag_ean, flag_cancel=flag_cancel)

        self.lbl_linhas.setText(f"Linhas: {len(df):,}".replace(",", "."))

        # Resumo Executivo: cards + gráficos + insights
        self.atualizar_exec(df)

        # Alertas
        self.atualizar_alertas(df)

        # Qualidade
        self.lbl_quality.setText(self.quality_text(df))

    def atualizar_exec(self, df: pd.DataFrame):
        k = self.calcular_kpis(df)

        # cards
        if not k:
            self.card_fat_liq.set_value("—")
            self.card_fat_bruto.set_value("—")
            self.card_ticket.set_value("—")
            self.card_vendas.set_value("—")
            self.card_itens.set_value("—")
            self.card_margem.set_value("—")
            self.card_impacto.set_value("—")
            self.txt_insights.setPlainText("Carregue dados para ver insights.")
        else:
            self.card_fat_liq.set_value(format_brl(k["fat_liq"]), "após excluir cancel/deneg")
            self.card_fat_bruto.set_value(format_brl(k["fat_bruto"]), "total bruto do recorte")
            self.card_ticket.set_value(format_brl(k["ticket"]), "média por venda")
            self.card_vendas.set_value(str(k["qtd_vendas"]), "vendas distintas")
            self.card_itens.set_value(str(k["qtd_itens"]), "itens de venda")
            if k["margem_pct"] is None:
                self.card_margem.set_value("—", "necessário custo_total")
            else:
                self.card_margem.set_value(f"{k['margem_pct']*100:.2f}%".replace(".", ","), "lucro bruto / faturamento")
            self.card_impacto.set_value(format_brl(k["impacto"]), "perdas por status inválido")

            self.txt_insights.setPlainText(self.insights_text(df))

        # gráficos no resumo executivo
        self.plot_exec_charts(df)

    def plot_exec_charts(self, df: pd.DataFrame):
        # Mensal
        self.fig_mes.clear()
        ax1 = self.fig_mes.add_subplot(111)
        ax1.set_title("Mensal", fontsize=10, fontweight="bold")
        serie = self.serie_mensal(df)
        if len(serie) > 0:
            x = [str(p) for p in serie.index]
            ax1.plot(x, serie.values, marker="o")
            ax1.tick_params(axis="x", rotation=45, labelsize=8)
        else:
            ax1.text(0.5, 0.5, "Sem data", ha="center", va="center")
        self.fig_mes.tight_layout()
        self.canvas_mes.draw()

        # Top produtos
        self.fig_top.clear()
        ax2 = self.fig_top.add_subplot(111)
        ax2.set_title("Top Produtos", fontsize=10, fontweight="bold")
        top = self.top_produtos(df)
        if len(top) > 0:
            ax2.bar(top.index.astype(str), top.values)
            ax2.tick_params(axis="x", rotation=45, labelsize=8)
        else:
            ax2.text(0.5, 0.5, "Sem produto", ha="center", va="center")
        self.fig_top.tight_layout()
        self.canvas_top.draw()

        # Canal
        self.fig_canal.clear()
        ax3 = self.fig_canal.add_subplot(111)
        ax3.set_title("Por Canal", fontsize=10, fontweight="bold")
        canal = self.por_canal(df)
        if len(canal) > 0:
            ax3.bar(canal.index.astype(str), canal.values)
            ax3.tick_params(axis="x", rotation=0, labelsize=8)
        else:
            ax3.text(0.5, 0.5, "Sem canal", ha="center", va="center")
        self.fig_canal.tight_layout()
        self.canvas_canal.draw()

    def atualizar_alertas(self, df: pd.DataFrame):
        alerts = self.gerar_alertas(df)
        self.model_alertas.update(alerts)

        if alerts.empty:
            self.lbl_alertas.setText("✅ Nenhum alerta relevante no recorte atual.")
            return

        resumo = alerts["tipo_alerta"].value_counts()
        texto = "🚨 Alertas (recorte atual)\n\n" + "\n".join([f"- {k}: {int(v)}" for k, v in resumo.items()])
        texto += "\n\nObs.: Linhas com alerta estão destacadas na aba Tabela."
        self.lbl_alertas.setText(texto)

    # =========================================================
    # Export Excel v6
    # =========================================================
    def on_exportar(self):
        if self.base.empty:
            return

        df = self.base_filtrada if not self.base_filtrada.empty else self.base

        arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Excel v6",
            str(self.base_dir / "relatorio_dashboard_v6.xlsx"),
            "Excel (*.xlsx)"
        )
        if not arquivo:
            return

        try:
            k = self.calcular_kpis(df)
            top = self.top_produtos(df).reset_index()
            if not top.empty:
                top.columns = ["produto", "faturamento"]

            serie = self.serie_mensal(df).reset_index()
            if not serie.empty:
                serie.columns = ["mes", "faturamento"]

            canal = self.por_canal(df).reset_index()
            if not canal.empty:
                canal.columns = ["canal", "faturamento"]

            alerts = self.gerar_alertas(df)
            quality = pd.DataFrame({"texto": [self.quality_text(df)]})
            insights = pd.DataFrame({"insights": [self.insights_text(df)]})

            resumo_exec = pd.DataFrame([{
                "faturamento_liquido": k.get("fat_liq", 0.0),
                "faturamento_bruto": k.get("fat_bruto", 0.0),
                "impacto_cancel_deneg": k.get("impacto", 0.0),
                "ticket_medio": k.get("ticket", 0.0),
                "qtd_vendas": k.get("qtd_vendas", 0),
                "qtd_itens": k.get("qtd_itens", 0),
                "margem_bruta_pct": (k["margem_pct"] if k.get("margem_pct") is not None else None)
            }])

            sheets = {
                "Resumo_Executivo": resumo_exec,
                "Base_Filtrada": df,
                "Top_Produtos": top,
                "Faturamento_Mensal": serie,
                "Faturamento_Canal": canal,
                "Alertas": alerts,
                "Qualidade": quality,
                "Insights": insights
            }

            with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:
                for name, dfx in sheets.items():
                    if dfx is None or len(dfx) == 0:
                        pd.DataFrame({"info": ["Sem dados para esta aba no recorte atual."]}).to_excel(
                            writer, sheet_name=name[:31], index=False
                        )
                    else:
                        dfx.to_excel(writer, sheet_name=name[:31], index=False)

            add_excel_format_and_charts(arquivo)

            QMessageBox.information(self, "Exportado", f"✅ Excel v6 salvo:\n{arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro ao exportar", str(e))


def main():
    app = QApplication(sys.argv)
    w = DashboardV6()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()