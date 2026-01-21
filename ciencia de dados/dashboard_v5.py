import sys
from pathlib import Path
from datetime import datetime

import pandas as pd

from PySide6.QtCore import Qt, QAbstractTableModel, QDate
from PySide6.QtGui import QColor
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFileDialog,
    QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QLineEdit, QTabWidget, QTableView, QMessageBox,
    QComboBox, QGroupBox, QFormLayout, QDateEdit
)

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference


# =========================================================
# Utils
# =========================================================
def format_brl(valor: float) -> str:
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    """
    Torna o dataset robusto (aceita variações de nomes).
    """
    base = df.copy()

    # Produto
    if "produto" not in base.columns and "descricao" in base.columns:
        base.rename(columns={"descricao": "produto"}, inplace=True)

    # Canal
    if "canal" not in base.columns and "canal_aquisicao" in base.columns:
        base.rename(columns={"canal_aquisicao": "canal"}, inplace=True)

    # Data
    if "data" not in base.columns and "dh_emissao" in base.columns:
        base.rename(columns={"dh_emissao": "data"}, inplace=True)

    if "data" in base.columns:
        base["data"] = pd.to_datetime(base["data"], errors="coerce")

    # Valores
    if "valor_total" not in base.columns and "v_prod" in base.columns:
        base.rename(columns={"v_prod": "valor_total"}, inplace=True)

    if "preco_unitario" not in base.columns and "v_un" in base.columns:
        base.rename(columns={"v_un": "preco_unitario"}, inplace=True)

    # EAN
    if "ean" in base.columns:
        base["ean"] = base["ean"].astype(str)

    # Nome cliente
    if "nome" not in base.columns:
        for cand in ["razao_social", "cliente", "destinatario_nome"]:
            if cand in base.columns:
                base.rename(columns={cand: "nome"}, inplace=True)
                break

    return base


def checar_arquivos_erp(data_dir: Path):
    arquivos = ["clientes.csv", "produtos.csv", "vendas.csv", "itens_venda.csv"]
    return [a for a in arquivos if not (data_dir / a).exists()]


def carregar_base_erp(data_dir: Path) -> pd.DataFrame:
    clientes = pd.read_csv(data_dir / "clientes.csv")
    produtos = pd.read_csv(data_dir / "produtos.csv")
    vendas = pd.read_csv(data_dir / "vendas.csv", parse_dates=["data"])
    itens = pd.read_csv(data_dir / "itens_venda.csv")

    base = (itens
            .merge(vendas, on="venda_id", how="left")
            .merge(produtos, on="produto_id", how="left")
            .merge(clientes, on="cliente_id", how="left"))

    base = normalize_cols(base)
    return base


def compute_outlier_mask(series: pd.Series) -> pd.Series:
    """Outlier por IQR (robusto)."""
    if series is None or series.empty:
        return pd.Series([False] * 0)

    s = pd.to_numeric(series, errors="coerce").dropna()
    if s.empty:
        return pd.Series([False] * len(series), index=series.index)

    q1 = s.quantile(0.25)
    q3 = s.quantile(0.75)
    iqr = q3 - q1
    lim_sup = q3 + 1.5 * iqr
    return pd.to_numeric(series, errors="coerce") > lim_sup


def add_excel_format_and_charts(path_xlsx: str):
    """
    Aplica formatação e adiciona gráficos no Excel.
    Espera algumas abas: KPIs, Faturamento_Mensal, Top_Produtos, Faturamento_Canal
    """
    import openpyxl

    wb = openpyxl.load_workbook(path_xlsx)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        max_col = ws.max_column
        max_row = ws.max_row

        # Cabeçalho
        for col in range(1, max_col + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center

        # Ajuste largura (simples)
        for col in range(1, max_col + 1):
            col_letter = get_column_letter(col)
            max_len = 10
            for row in range(1, min(max_row, 200) + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                max_len = max(max_len, min(len(str(v)), 50))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 42)

        # Formato moeda por heurística
        for col in range(1, max_col + 1):
            header = str(ws.cell(row=1, column=col).value).lower()
            if any(k in header for k in ["valor", "fatur", "custo", "preco", "total", "lucro", "impacto"]):
                for row in range(2, max_row + 1):
                    c = ws.cell(row=row, column=col)
                    if isinstance(c.value, (int, float)):
                        c.number_format = '"R$" #,##0.00'

    # =========== CHARTS ===========
    # Faturamento Mensal (line)
    if "Faturamento_Mensal" in wb.sheetnames:
        ws = wb["Faturamento_Mensal"]
        if ws.max_row >= 2 and ws.max_column >= 2:
            chart = LineChart()
            chart.title = "Faturamento Mensal"
            chart.y_axis.title = "R$"
            chart.x_axis.title = "Mês"

            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 24
            ws.add_chart(chart, "D2")

    # Top Produtos (bar)
    if "Top_Produtos" in wb.sheetnames:
        ws = wb["Top_Produtos"]
        if ws.max_row >= 2 and ws.max_column >= 2:
            chart = BarChart()
            chart.title = "Top Produtos"
            chart.y_axis.title = "R$"
            chart.x_axis.title = "Produto"

            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 24
            ws.add_chart(chart, "D2")

    # Faturamento por Canal (bar)
    if "Faturamento_Canal" in wb.sheetnames:
        ws = wb["Faturamento_Canal"]
        if ws.max_row >= 2 and ws.max_column >= 2:
            chart = BarChart()
            chart.title = "Faturamento por Canal"
            chart.y_axis.title = "R$"
            chart.x_axis.title = "Canal"

            data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
            cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)

            chart.add_data(data, titles_from_data=True)
            chart.set_categories(cats)
            chart.height = 10
            chart.width = 24
            ws.add_chart(chart, "D2")

    wb.save(path_xlsx)


# =========================================================
# Table Model com destaque de alertas
# =========================================================
class PandasTableModel(QAbstractTableModel):
    def __init__(self, df: pd.DataFrame):
        super().__init__()
        self.df = df.copy()
        self.flag_outlier = pd.Series([False] * len(df))
        self.flag_ean_invalid = pd.Series([False] * len(df))
        self.flag_cancel = pd.Series([False] * len(df))

    def rowCount(self, parent=None):
        return len(self.df)

    def columnCount(self, parent=None):
        return len(self.df.columns)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None

        r = index.row()
        c = index.column()

        if role == Qt.DisplayRole:
            value = self.df.iloc[r, c]
            if pd.isna(value):
                return ""

            if isinstance(value, float):
                return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            return str(value)

        # Destaque visual (linhas suspeitas)
        if role == Qt.BackgroundRole:
            try:
                if len(self.flag_cancel) > 0 and bool(self.flag_cancel.iloc[r]):
                    return QColor(255, 220, 220)  # vermelho claro
                if len(self.flag_outlier) > 0 and bool(self.flag_outlier.iloc[r]):
                    return QColor(255, 245, 204)  # amarelo claro
                if len(self.flag_ean_invalid) > 0 and bool(self.flag_ean_invalid.iloc[r]):
                    return QColor(220, 235, 255)  # azul claro
            except Exception:
                return None

        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return str(self.df.columns[section])
        return str(section)

    def update(self, df: pd.DataFrame, flag_outlier=None, flag_ean_invalid=None, flag_cancel=None):
        self.beginResetModel()
        self.df = df.copy()

        n = len(df)
        self.flag_outlier = flag_outlier if flag_outlier is not None else pd.Series([False] * n)
        self.flag_ean_invalid = flag_ean_invalid if flag_ean_invalid is not None else pd.Series([False] * n)
        self.flag_cancel = flag_cancel if flag_cancel is not None else pd.Series([False] * n)

        # garantir index alinhado
        self.flag_outlier = self.flag_outlier.reindex(df.index, fill_value=False)
        self.flag_ean_invalid = self.flag_ean_invalid.reindex(df.index, fill_value=False)
        self.flag_cancel = self.flag_cancel.reindex(df.index, fill_value=False)

        self.endResetModel()


# =========================================================
# Dashboard v5
# =========================================================
class DashboardV5(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Dashboard v5 - Ciência de Dados (ERP/Fiscal) | Portfólio Premium")
        self.resize(1450, 820)

        self.base_dir = Path(__file__).resolve().parent
        self.data_dir = self.base_dir / "dados"

        self.base = pd.DataFrame()
        self.base_filtrada = pd.DataFrame()

        # -------------------------
        # Layout principal
        # -------------------------
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)

        # Barra superior
        top_bar = QHBoxLayout()
        self.btn_carregar = QPushButton("Carregar (pasta dados/)")
        self.btn_carregar.clicked.connect(self.on_carregar)

        self.btn_escolher = QPushButton("Escolher pasta...")
        self.btn_escolher.clicked.connect(self.on_escolher_pasta)

        self.btn_limpar = QPushButton("Limpar filtros")
        self.btn_limpar.clicked.connect(self.on_limpar)
        self.btn_limpar.setEnabled(False)

        self.btn_exportar = QPushButton("Exportar Excel v5 (com gráficos)")
        self.btn_exportar.clicked.connect(self.on_exportar)
        self.btn_exportar.setEnabled(False)

        self.lbl_status = QLabel("Status: aguardando...")
        self.lbl_status.setStyleSheet("font-weight:bold;")

        top_bar.addWidget(self.btn_carregar)
        top_bar.addWidget(self.btn_escolher)
        top_bar.addWidget(self.btn_limpar)
        top_bar.addWidget(self.btn_exportar)
        top_bar.addStretch()
        top_bar.addWidget(self.lbl_status)
        main_layout.addLayout(top_bar)

        # -------------------------
        # Filtros
        # -------------------------
        filtros_box = QGroupBox("Filtros (Texto + UF + Canal + Categoria + Status + Período)")
        filtros_layout = QHBoxLayout(filtros_box)

        self.input_busca = QLineEdit()
        self.input_busca.setPlaceholderText("Buscar (produto, EAN, UF, canal, cliente, CFOP, CST...)")
        self.input_busca.textChanged.connect(self.aplicar_filtros)

        self.combo_uf = QComboBox()
        self.combo_uf.currentIndexChanged.connect(self.aplicar_filtros)

        self.combo_canal = QComboBox()
        self.combo_canal.currentIndexChanged.connect(self.aplicar_filtros)

        self.combo_categoria = QComboBox()
        self.combo_categoria.currentIndexChanged.connect(self.aplicar_filtros)

        self.combo_status = QComboBox()
        self.combo_status.currentIndexChanged.connect(self.aplicar_filtros)

        # Filtro de período
        self.date_ini = QDateEdit()
        self.date_ini.setCalendarPopup(True)
        self.date_ini.dateChanged.connect(self.aplicar_filtros)

        self.date_fim = QDateEdit()
        self.date_fim.setCalendarPopup(True)
        self.date_fim.dateChanged.connect(self.aplicar_filtros)

        self.lbl_linhas = QLabel("Linhas: 0")
        self.lbl_linhas.setStyleSheet("font-weight:bold;")

        filtros_layout.addWidget(QLabel("Texto:"))
        filtros_layout.addWidget(self.input_busca, 2)
        filtros_layout.addWidget(self.combo_uf)
        filtros_layout.addWidget(self.combo_canal)
        filtros_layout.addWidget(self.combo_categoria)
        filtros_layout.addWidget(self.combo_status)
        filtros_layout.addWidget(QLabel("De:"))
        filtros_layout.addWidget(self.date_ini)
        filtros_layout.addWidget(QLabel("Até:"))
        filtros_layout.addWidget(self.date_fim)
        filtros_layout.addWidget(self.lbl_linhas)

        main_layout.addWidget(filtros_box)

        # -------------------------
        # Tabs
        # -------------------------
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        # TAB 1 - Tabela
        self.tab_tabela = QWidget()
        self.tabs.addTab(self.tab_tabela, "Tabela")
        t1 = QVBoxLayout(self.tab_tabela)
        self.table_base = QTableView()
        self.table_base.setSortingEnabled(True)
        self.model_base = PandasTableModel(pd.DataFrame())
        self.table_base.setModel(self.model_base)
        t1.addWidget(self.table_base)

        # TAB 2 - KPIs
        self.tab_kpis = QWidget()
        self.tabs.addTab(self.tab_kpis, "KPIs")
        t2 = QHBoxLayout(self.tab_kpis)

        kpi_box = QGroupBox("KPIs (dinâmicos)")
        kpi_form = QFormLayout(kpi_box)

        self.lbl_fat_bruto = QLabel("-")
        self.lbl_fat_liq = QLabel("-")
        self.lbl_impacto = QLabel("-")
        self.lbl_ticket = QLabel("-")
        self.lbl_vendas = QLabel("-")
        self.lbl_itens = QLabel("-")
        self.lbl_margem = QLabel("-")

        kpi_form.addRow("Faturamento bruto:", self.lbl_fat_bruto)
        kpi_form.addRow("Faturamento líquido:", self.lbl_fat_liq)
        kpi_form.addRow("Impacto cancel/deneg:", self.lbl_impacto)
        kpi_form.addRow("Ticket médio:", self.lbl_ticket)
        kpi_form.addRow("Qtd vendas:", self.lbl_vendas)
        kpi_form.addRow("Qtd itens:", self.lbl_itens)
        kpi_form.addRow("Margem bruta:", self.lbl_margem)

        t2.addWidget(kpi_box, 1)

        # TAB 3 - Gráficos (sub-tabs)
        self.tab_graficos = QWidget()
        self.tabs.addTab(self.tab_graficos, "Gráficos")
        t3 = QVBoxLayout(self.tab_graficos)
        self.tabs_graf = QTabWidget()
        t3.addWidget(self.tabs_graf)

        # Top Produtos
        self.tab_top = QWidget()
        self.tabs_graf.addTab(self.tab_top, "Top Produtos")
        lt = QVBoxLayout(self.tab_top)
        self.fig_top = Figure(figsize=(6, 4))
        self.canvas_top = FigureCanvas(self.fig_top)
        lt.addWidget(self.canvas_top)

        # Mensal
        self.tab_mes = QWidget()
        self.tabs_graf.addTab(self.tab_mes, "Faturamento Mensal")
        lm = QVBoxLayout(self.tab_mes)
        self.fig_mes = Figure(figsize=(6, 4))
        self.canvas_mes = FigureCanvas(self.fig_mes)
        lm.addWidget(self.canvas_mes)

        # Canal
        self.tab_canal = QWidget()
        self.tabs_graf.addTab(self.tab_canal, "Por Canal")
        lc = QVBoxLayout(self.tab_canal)
        self.fig_canal = Figure(figsize=(6, 4))
        self.canvas_canal = FigureCanvas(self.fig_canal)
        lc.addWidget(self.canvas_canal)

        # TAB 4 - Alertas
        self.tab_alertas = QWidget()
        self.tabs.addTab(self.tab_alertas, "Alertas")
        ta = QVBoxLayout(self.tab_alertas)

        self.lbl_alertas = QLabel("Carregue os dados para ver os alertas.")
        self.lbl_alertas.setTextInteractionFlags(Qt.TextSelectableByMouse)
        ta.addWidget(self.lbl_alertas)

        self.table_alertas = QTableView()
        self.table_alertas.setSortingEnabled(True)
        self.model_alertas = PandasTableModel(pd.DataFrame())
        self.table_alertas.setModel(self.model_alertas)
        ta.addWidget(self.table_alertas)

        # TAB 5 - Qualidade
        self.tab_quality = QWidget()
        self.tabs.addTab(self.tab_quality, "Qualidade de Dados")
        tq = QVBoxLayout(self.tab_quality)

        self.lbl_quality = QLabel("Carregue os dados para ver a auditoria.")
        self.lbl_quality.setTextInteractionFlags(Qt.TextSelectableByMouse)
        tq.addWidget(self.lbl_quality)

        # Autoload
        if self.data_dir.exists():
            self.on_carregar()

    # =========================================================
    # Carregamento
    # =========================================================
    def on_escolher_pasta(self):
        folder = QFileDialog.getExistingDirectory(self, "Selecione a pasta dos CSVs")
        if folder:
            self.data_dir = Path(folder)
            self.on_carregar()

    def on_carregar(self):
        try:
            faltando = checar_arquivos_erp(self.data_dir)
            if faltando:
                raise FileNotFoundError(f"Arquivos faltando em {self.data_dir}: {faltando}")

            self.base = carregar_base_erp(self.data_dir)
            self.base_filtrada = self.base.copy()

            self.preencher_combos()
            self.preencher_periodo()

            self.atualizar_tudo()

            self.btn_limpar.setEnabled(True)
            self.btn_exportar.setEnabled(True)
            self.lbl_status.setText(f"Status: OK | Base: {self.base.shape[0]} linhas / {self.base.shape[1]} colunas")

        except Exception as e:
            QMessageBox.critical(self, "Erro ao carregar", str(e))
            self.lbl_status.setText("Status: ERRO ao carregar")

    def preencher_combos(self):
        self.combo_uf.blockSignals(True)
        self.combo_canal.blockSignals(True)
        self.combo_categoria.blockSignals(True)
        self.combo_status.blockSignals(True)

        self.combo_uf.clear()
        self.combo_canal.clear()
        self.combo_categoria.clear()
        self.combo_status.clear()

        self.combo_uf.addItem("UF: Todas")
        self.combo_canal.addItem("Canal: Todos")
        self.combo_categoria.addItem("Categoria: Todas")
        self.combo_status.addItem("Status: Todos")

        if "uf_destino" in self.base.columns:
            for u in sorted(self.base["uf_destino"].dropna().unique().tolist()):
                self.combo_uf.addItem(f"UF: {u}")

        if "canal" in self.base.columns:
            for c in sorted(self.base["canal"].dropna().unique().tolist()):
                self.combo_canal.addItem(f"Canal: {c}")

        if "categoria" in self.base.columns:
            for cat in sorted(self.base["categoria"].dropna().unique().tolist()):
                self.combo_categoria.addItem(f"Categoria: {cat}")

        if "status" in self.base.columns:
            for st in sorted(self.base["status"].dropna().unique().tolist()):
                self.combo_status.addItem(f"Status: {st}")

        self.combo_uf.blockSignals(False)
        self.combo_canal.blockSignals(False)
        self.combo_categoria.blockSignals(False)
        self.combo_status.blockSignals(False)

    def preencher_periodo(self):
        """Define De/Até baseado no dataset."""
        if "data" not in self.base.columns:
            hoje = QDate.currentDate()
            self.date_ini.setDate(hoje.addMonths(-6))
            self.date_fim.setDate(hoje)
            return

        d = self.base["data"].dropna()
        if d.empty:
            hoje = QDate.currentDate()
            self.date_ini.setDate(hoje.addMonths(-6))
            self.date_fim.setDate(hoje)
            return

        min_dt = d.min().date()
        max_dt = d.max().date()

        self.date_ini.blockSignals(True)
        self.date_fim.blockSignals(True)

        self.date_ini.setDate(QDate(min_dt.year, min_dt.month, min_dt.day))
        self.date_fim.setDate(QDate(max_dt.year, max_dt.month, max_dt.day))

        self.date_ini.blockSignals(False)
        self.date_fim.blockSignals(False)

    # =========================================================
    # Filtros
    # =========================================================
    def on_limpar(self):
        self.input_busca.setText("")
        self.combo_uf.setCurrentIndex(0)
        self.combo_canal.setCurrentIndex(0)
        self.combo_categoria.setCurrentIndex(0)
        self.combo_status.setCurrentIndex(0)

        # reset período para intervalo total
        self.preencher_periodo()

        self.base_filtrada = self.base.copy()
        self.atualizar_tudo()

    def aplicar_filtros(self):
        if self.base.empty:
            return

        df = self.base.copy()

        # Período
        if "data" in df.columns:
            di = self.date_ini.date().toPython()
            dfim = self.date_fim.date().toPython()

            df["data"] = pd.to_datetime(df["data"], errors="coerce")
            df = df.dropna(subset=["data"])

            df = df[(df["data"].dt.date >= di) & (df["data"].dt.date <= dfim)]

        # Texto
        termo = self.input_busca.text().strip().lower()
        if termo:
            cols_preferidas = [c for c in df.columns if c.lower() in [
                "produto", "ean", "uf_destino", "canal", "categoria",
                "cfop", "cst_icms", "nome", "marca"
            ]]
            if not cols_preferidas:
                cols_preferidas = df.columns.tolist()

            mask = df[cols_preferidas].astype(str).apply(
                lambda col: col.str.lower().str.contains(termo, na=False)
            ).any(axis=1)
            df = df[mask]

        # UF
        uf_sel = self.combo_uf.currentText().replace("UF: ", "")
        if uf_sel != "Todas" and "uf_destino" in df.columns:
            df = df[df["uf_destino"] == uf_sel]

        # Canal
        canal_sel = self.combo_canal.currentText().replace("Canal: ", "")
        if canal_sel != "Todos" and "canal" in df.columns:
            df = df[df["canal"] == canal_sel]

        # Categoria
        cat_sel = self.combo_categoria.currentText().replace("Categoria: ", "")
        if cat_sel != "Todas" and "categoria" in df.columns:
            df = df[df["categoria"] == cat_sel]

        # Status
        st_sel = self.combo_status.currentText().replace("Status: ", "")
        if st_sel != "Todos" and "status" in df.columns:
            df = df[df["status"] == st_sel]

        self.base_filtrada = df.copy()
        self.atualizar_tudo()

    # =========================================================
    # KPIs
    # =========================================================
    def faturamento_liquido(self, df: pd.DataFrame) -> tuple[float, float, float]:
        if "valor_total" not in df.columns:
            return 0.0, 0.0, 0.0

        fat_bruto = float(df["valor_total"].sum())

        if "status" not in df.columns:
            return fat_bruto, fat_bruto, 0.0

        status = df["status"].astype(str).str.upper()
        invalidos = status.str.contains("CANCEL") | status.str.contains("DENEG") | status.str.contains("INUTIL")

        df_ok = df[~invalidos].copy()
        fat_liq = float(df_ok["valor_total"].sum())
        impacto = fat_bruto - fat_liq
        return fat_bruto, fat_liq, impacto

    def calcular_kpis(self, df: pd.DataFrame) -> dict:
        if df.empty or "valor_total" not in df.columns:
            return {}

        fat_bruto, fat_liq, impacto = self.faturamento_liquido(df)

        qtd_vendas = int(df["venda_id"].nunique()) if "venda_id" in df.columns else int(df.shape[0])
        ticket = float(df.groupby("venda_id")["valor_total"].sum().mean()) if "venda_id" in df.columns else fat_bruto
        qtd_itens = int(df.shape[0])

        margem_pct = None
        if "custo_total" in df.columns:
            custo = float(df["custo_total"].sum())
            lucro = fat_bruto - custo
            margem_pct = float(lucro / fat_bruto) if fat_bruto != 0 else 0.0

        return {
            "fat_bruto": fat_bruto,
            "fat_liq": fat_liq,
            "impacto": impacto,
            "qtd_vendas": qtd_vendas,
            "ticket": ticket,
            "qtd_itens": qtd_itens,
            "margem_pct": margem_pct
        }

    # =========================================================
    # Agregações
    # =========================================================
    def top_produtos(self, df: pd.DataFrame) -> pd.Series:
        if "produto" not in df.columns or "valor_total" not in df.columns:
            return pd.Series(dtype=float)
        return (df.groupby("produto")["valor_total"].sum().sort_values(ascending=False).head(10))

    def serie_mensal(self, df: pd.DataFrame) -> pd.Series:
        if "data" not in df.columns or "valor_total" not in df.columns:
            return pd.Series(dtype=float)
        tmp = df.dropna(subset=["data"]).copy()
        if tmp.empty:
            return pd.Series(dtype=float)
        return (tmp.groupby(tmp["data"].dt.to_period("M"))["valor_total"].sum().sort_index())

    def por_canal(self, df: pd.DataFrame) -> pd.Series:
        if "canal" not in df.columns or "valor_total" not in df.columns:
            return pd.Series(dtype=float)
        return (df.groupby("canal")["valor_total"].sum().sort_values(ascending=False))

    # =========================================================
    # Alertas (linhas suspeitas)
    # =========================================================
    def gerar_alertas(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Retorna DataFrame com "tipo_alerta" + colunas úteis (amostra de problemas).
        """
        if df.empty:
            return pd.DataFrame()

        alerts = []

        # Canceladas/Denegadas
        if "status" in df.columns:
            st = df["status"].astype(str).str.upper()
            mask_cancel = st.str.contains("CANCEL") | st.str.contains("DENEG") | st.str.contains("INUTIL")
            if mask_cancel.any():
                a = df[mask_cancel].copy()
                a["tipo_alerta"] = "STATUS_CANCEL_DENEG"
                alerts.append(a)

        # EAN inválido
        if "ean" in df.columns:
            ean_len = df["ean"].astype(str).str.len()
            mask_ean = ean_len != 13
            if mask_ean.any():
                a = df[mask_ean].copy()
                a["tipo_alerta"] = "EAN_INVALIDO_LEN"
                alerts.append(a)

            # EAN duplicado (por item)
            dup = df[df.duplicated(subset=["ean"], keep=False)].copy()
            if not dup.empty:
                dup["tipo_alerta"] = "EAN_DUPLICADO"
                alerts.append(dup)

        # Outlier de preco_unitario
        if "preco_unitario" in df.columns:
            mask_out = compute_outlier_mask(df["preco_unitario"])
            if mask_out.any():
                a = df[mask_out].copy()
                a["tipo_alerta"] = "OUTLIER_PRECO_UNITARIO"
                alerts.append(a)

        if not alerts:
            return pd.DataFrame()

        final = pd.concat(alerts, ignore_index=True)

        # Mostrar só colunas relevantes se existirem
        cols_show = []
        for c in ["tipo_alerta", "data", "venda_id", "produto_id", "produto", "ean",
                  "quantidade", "preco_unitario", "valor_total", "uf_destino", "cfop", "status", "nome"]:
            if c in final.columns:
                cols_show.append(c)

        if cols_show:
            final = final[cols_show]

        return final.head(800)  # limite para não travar

    def gerar_flags_tabela(self, df: pd.DataFrame):
        """Flags para destacar linhas na tabela principal."""
        n = len(df)
        if n == 0:
            return pd.Series([], dtype=bool), pd.Series([], dtype=bool), pd.Series([], dtype=bool)

        # Outlier preco_unitario
        flag_out = pd.Series([False] * n, index=df.index)
        if "preco_unitario" in df.columns:
            flag_out = compute_outlier_mask(df["preco_unitario"]).fillna(False)

        # EAN inválido
        flag_ean = pd.Series([False] * n, index=df.index)
        if "ean" in df.columns:
            flag_ean = df["ean"].astype(str).str.len().ne(13)

        # Canceladas/denegadas
        flag_cancel = pd.Series([False] * n, index=df.index)
        if "status" in df.columns:
            st = df["status"].astype(str).str.upper()
            flag_cancel = st.str.contains("CANCEL") | st.str.contains("DENEG") | st.str.contains("INUTIL")

        return flag_out, flag_ean, flag_cancel

    # =========================================================
    # Qualidade (texto)
    # =========================================================
    def quality_text(self, df: pd.DataFrame) -> str:
        if df.empty:
            return "🧪 Qualidade de Dados\n\nSem dados no recorte atual."

        linhas = []
        linhas.append(f"- Linhas: {df.shape[0]}")
        linhas.append(f"- Colunas: {df.shape[1]}")

        nulos = df.isna().sum().sort_values(ascending=False).head(15)
        linhas.append("\nTop 15 colunas com nulos:")
        for col, qtd in nulos.items():
            if qtd > 0:
                linhas.append(f"  • {col}: {int(qtd)}")

        if "ean" in df.columns:
            inv = int(df["ean"].astype(str).str.len().ne(13).sum())
            dup = int(df[df.duplicated(subset=["ean"], keep=False)]["ean"].nunique())
            linhas.append(f"\nEAN inválidos (len != 13): {inv}")
            linhas.append(f"EAN duplicados (distintos): {dup}")

        if "preco_unitario" in df.columns:
            out = int(compute_outlier_mask(df["preco_unitario"]).sum())
            linhas.append(f"\nOutliers preco_unitario (IQR): {out}")

        return "🧪 Qualidade de Dados (recorte atual)\n\n" + "\n".join(linhas)

    # =========================================================
    # Atualizações UI
    # =========================================================
    def atualizar_tudo(self):
        df = self.base_filtrada if not self.base_filtrada.empty else self.base

        # flags para destaque na tabela
        flag_out, flag_ean, flag_cancel = self.gerar_flags_tabela(df)

        # tabela base
        self.model_base.update(df, flag_outlier=flag_out, flag_ean_invalid=flag_ean, flag_cancel=flag_cancel)
        self.lbl_linhas.setText(f"Linhas: {len(df):,}".replace(",", "."))

        # KPIs
        self.atualizar_kpis(df)

        # gráficos
        self.atualizar_graficos(df)

        # Alertas
        self.atualizar_alertas(df)

        # Qualidade
        self.lbl_quality.setText(self.quality_text(df))

    def atualizar_kpis(self, df: pd.DataFrame):
        k = self.calcular_kpis(df)
        if not k:
            self.lbl_fat_bruto.setText("—")
            self.lbl_fat_liq.setText("—")
            self.lbl_impacto.setText("—")
            self.lbl_ticket.setText("—")
            self.lbl_vendas.setText("—")
            self.lbl_itens.setText("—")
            self.lbl_margem.setText("—")
            return

        self.lbl_fat_bruto.setText(format_brl(k["fat_bruto"]))
        self.lbl_fat_liq.setText(format_brl(k["fat_liq"]))
        self.lbl_impacto.setText(format_brl(k["impacto"]))
        self.lbl_ticket.setText(format_brl(k["ticket"]))
        self.lbl_vendas.setText(str(k["qtd_vendas"]))
        self.lbl_itens.setText(str(k["qtd_itens"]))

        if k["margem_pct"] is None:
            self.lbl_margem.setText("—")
        else:
            self.lbl_margem.setText(f"{k['margem_pct']*100:.2f}%".replace(".", ","))

    def atualizar_graficos(self, df: pd.DataFrame):
        # Top Produtos
        top = self.top_produtos(df)
        self.fig_top.clear()
        ax = self.fig_top.add_subplot(111)
        ax.set_title("Top 10 Produtos por Faturamento")
        if len(top) > 0:
            ax.bar(top.index.astype(str), top.values)
            ax.tick_params(axis="x", rotation=45)
        else:
            ax.text(0.5, 0.5, "Sem coluna 'produto' no recorte.", ha="center", va="center")
        self.fig_top.tight_layout()
        self.canvas_top.draw()

        # Mensal
        serie = self.serie_mensal(df)
        self.fig_mes.clear()
        ax2 = self.fig_mes.add_subplot(111)
        ax2.set_title("Faturamento Mensal")
        if len(serie) > 0:
            x = [str(p) for p in serie.index]
            ax2.plot(x, serie.values, marker="o")
            ax2.tick_params(axis="x", rotation=45)
        else:
            ax2.text(0.5, 0.5, "Sem coluna 'data' no recorte.", ha="center", va="center")
        self.fig_mes.tight_layout()
        self.canvas_mes.draw()

        # Canal
        canal = self.por_canal(df)
        self.fig_canal.clear()
        ax3 = self.fig_canal.add_subplot(111)
        ax3.set_title("Faturamento por Canal")
        if len(canal) > 0:
            ax3.bar(canal.index.astype(str), canal.values)
        else:
            ax3.text(0.5, 0.5, "Sem coluna 'canal' no recorte.", ha="center", va="center")
        self.fig_canal.tight_layout()
        self.canvas_canal.draw()

    def atualizar_alertas(self, df: pd.DataFrame):
        alerts = self.gerar_alertas(df)
        self.model_alertas.update(alerts)

        # Resumo
        if alerts.empty:
            self.lbl_alertas.setText("✅ Nenhum alerta relevante no recorte atual.")
            return

        resumo = alerts["tipo_alerta"].value_counts()
        texto = "🚨 Alertas (amostra)\n\n" + "\n".join([f"- {k}: {v}" for k, v in resumo.items()])
        texto += "\n\nObservação: as linhas com alerta estão destacadas na aba Tabela."
        self.lbl_alertas.setText(texto)

    # =========================================================
    # Export Excel v5
    # =========================================================
    def on_exportar(self):
        if self.base.empty:
            return

        df = self.base_filtrada if not self.base_filtrada.empty else self.base

        arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Excel v5",
            str(self.base_dir / "relatorio_dashboard_v5.xlsx"),
            "Excel (*.xlsx)"
        )
        if not arquivo:
            return

        try:
            k = self.calcular_kpis(df)

            # Sheets
            top = self.top_produtos(df).reset_index()
            if not top.empty:
                top.columns = ["produto", "faturamento"]

            serie = self.serie_mensal(df).reset_index()
            if not serie.empty:
                serie.columns = ["mes", "faturamento"]

            canal = self.por_canal(df).reset_index()
            if not canal.empty:
                canal.columns = ["canal", "faturamento"]

            alerts = self.gerar_alertas(df)

            df_kpis = pd.DataFrame([{
                "faturamento_bruto": k.get("fat_bruto", 0.0),
                "faturamento_liquido": k.get("fat_liq", 0.0),
                "impacto_cancel_deneg": k.get("impacto", 0.0),
                "ticket_medio": k.get("ticket", 0.0),
                "qtd_vendas": k.get("qtd_vendas", 0),
                "qtd_itens": k.get("qtd_itens", 0),
                "margem_bruta_pct": (k["margem_pct"] if k.get("margem_pct") is not None else None)
            }])

            df_quality = pd.DataFrame({"texto": [self.quality_text(df)]})

            sheets = {
                "Base_Filtrada": df,
                "KPIs": df_kpis,
                "Top_Produtos": top,
                "Faturamento_Mensal": serie,
                "Faturamento_Canal": canal,
                "Alertas": alerts,
                "Qualidade": df_quality,
            }

            # Export simples
            with pd.ExcelWriter(arquivo, engine="openpyxl") as writer:
                for name, dfx in sheets.items():
                    if dfx is None or len(dfx) == 0:
                        pd.DataFrame({"info": ["Sem dados para esta aba no recorte atual."]}).to_excel(
                            writer, sheet_name=name[:31], index=False
                        )
                    else:
                        dfx.to_excel(writer, sheet_name=name[:31], index=False)

            # Formatar e inserir gráficos
            add_excel_format_and_charts(arquivo)

            QMessageBox.information(self, "Exportado", f"✅ Excel v5 com gráficos salvo:\n{arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro ao exportar", str(e))


def main():
    app = QApplication(sys.argv)
    w = DashboardV5()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()