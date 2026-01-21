import sys
from pathlib import Path
from datetime import datetime

import pandas as pd

from PySide6.QtCore import Qt, QAbstractTableModel
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QFileDialog,
    QVBoxLayout, QHBoxLayout, QPushButton, QLabel,
    QLineEdit, QTabWidget, QTableView, QMessageBox,
    QComboBox, QGroupBox, QFormLayout
)

from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.figure import Figure

from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# =========================================================
# Utils
# =========================================================
def format_brl(valor: float) -> str:
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def normalize_cols(base: pd.DataFrame) -> pd.DataFrame:
    """
    Deixa o dataset robusto para nomes diferentes de colunas.
    Ex.: canal_aquisicao -> canal
    """
    df = base.copy()

    # Produto
    if "produto" not in df.columns and "descricao" in df.columns:
        df.rename(columns={"descricao": "produto"}, inplace=True)

    # Canal
    if "canal" not in df.columns and "canal_aquisicao" in df.columns:
        df.rename(columns={"canal_aquisicao": "canal"}, inplace=True)

    # Data
    if "data" not in df.columns and "dh_emissao" in df.columns:
        df.rename(columns={"dh_emissao": "data"}, inplace=True)

    if "data" in df.columns:
        df["data"] = pd.to_datetime(df["data"], errors="coerce")

    # Valores e custos
    if "valor_total" not in df.columns and "v_prod" in df.columns:
        df.rename(columns={"v_prod": "valor_total"}, inplace=True)

    if "preco_unitario" not in df.columns and "v_un" in df.columns:
        df.rename(columns={"v_un": "preco_unitario"}, inplace=True)

    # EAN
    if "ean" in df.columns:
        df["ean"] = df["ean"].astype(str)

    # Cliente "nome"
    # (se não existir, tenta usar "razao_social" ou algo similar)
    if "nome" not in df.columns:
        for cand in ["razao_social", "cliente", "destinatario_nome"]:
            if cand in df.columns:
                df.rename(columns={cand: "nome"}, inplace=True)
                break

    return df


def export_excel_formatado(path_xlsx: str, sheets_dict: dict):
    """
    Cria Excel com múltiplas abas e aplica uma formatação simples e profissional.
    sheets_dict: { "SheetName": dataframe }
    """
    with pd.ExcelWriter(path_xlsx, engine="openpyxl") as writer:
        for sheet_name, df in sheets_dict.items():
            if df is None or len(df) == 0:
                # criar uma aba vazia com observação
                pd.DataFrame({"info": ["Sem dados para este recorte."]}).to_excel(
                    writer, sheet_name=sheet_name[:31], index=False
                )
            else:
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

        wb = writer.book

        # Formatar cada aba
        header_fill = PatternFill("solid", fgColor="1F4E79")  # azul escuro
        header_font = Font(color="FFFFFF", bold=True)
        center = Alignment(horizontal="center", vertical="center")

        for ws in wb.worksheets:
            ws.freeze_panes = "A2"

            # Ajustar largura
            max_col = ws.max_column
            max_row = ws.max_row

            # Cabeçalho
            for col in range(1, max_col + 1):
                cell = ws.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center

            # Ajustar largura por conteúdo (simples)
            for col in range(1, max_col + 1):
                col_letter = get_column_letter(col)
                max_len = 10
                for row in range(1, min(max_row, 200) + 1):
                    v = ws.cell(row=row, column=col).value
                    if v is None:
                        continue
                    max_len = max(max_len, len(str(v))[:50].__len__())
                ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

            # Formatos numéricos (heurística)
            for col in range(1, max_col + 1):
                header = str(ws.cell(row=1, column=col).value).lower()
                if any(k in header for k in ["valor", "faturamento", "custo", "preco", "total", "bruto", "liquido"]):
                    for row in range(2, max_row + 1):
                        c = ws.cell(row=row, column=col)
                        if isinstance(c.value, (int, float)):
                            c.number_format = '"R$" #,##0.00'
                elif any(k in header for k in ["aliq", "percent", "margem"]):
                    for row in range(2, max_row + 1):
                        c = ws.cell(row=row, column=col)
                        if isinstance(c.value, (int, float)):
                            c.number_format = '0.00%'

        wb.save(path_xlsx)


# =========================================================
# QTable Model
# =========================================================
class PandasTableModel(QAbstractTableModel):
    def __init__(self, df: pd.DataFrame):
        super().__init__()
        self.df = df.copy()

    def rowCount(self, parent=None):
        return len(self.df)

    def columnCount(self, parent=None):
        return len(self.df.columns)

    def data(self, index, role=Qt.DisplayRole):
        if not index.isValid():
            return None

        if role == Qt.DisplayRole:
            value = self.df.iloc[index.row(), index.column()]
            if pd.isna(value):
                return ""

            if isinstance(value, float):
                return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            return str(value)

        return None

    def headerData(self, section, orientation, role=Qt.DisplayRole):
        if role != Qt.DisplayRole:
            return None
        if orientation == Qt.Horizontal:
            return str(self.df.columns[section])
        return str(section)

    def update(self, df: pd.DataFrame):
        self.beginResetModel()
        self.df = df.copy()
        self.endResetModel()


# =========================================================
# Loader ERP
# =========================================================
def checar_arquivos_erp(data_dir: Path):
    arquivos = ["clientes.csv", "produtos.csv", "vendas.csv", "itens_venda.csv"]
    return [a for a in arquivos if not (data_dir / a).exists()]


def carregar_base_erp(data_dir: Path) -> pd.DataFrame:
    clientes = pd.read_csv(data_dir / "clientes.csv")
    produtos = pd.read_csv(data_dir / "produtos.csv")
    vendas = pd.read_csv(data_dir / "vendas.csv", parse_dates=["data"])
    itens = pd.read_csv(data_dir / "itens_venda.csv")

    base = (itens
            .merge(vendas, on="venda_id", how="left")
            .merge(produtos, on="produto_id", how="left")
            .merge(clientes, on="cliente_id", how="left"))

    base = normalize_cols(base)
    return base


# =========================================================
# Dashboard V4
# =========================================================
class DashboardV4(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Dashboard v4 - Ciência de Dados (ERP/Fiscal) | Portfólio")
        self.resize(1400, 800)

        self.base_dir = Path(__file__).resolve().parent
        self.data_dir = self.base_dir / "dados"

        self.base = pd.DataFrame()
        self.base_filtrada = pd.DataFrame()

        # -------------------------
        # Layout principal
        # -------------------------
        central = QWidget()
        self.setCentralWidget(central)
        main_layout = QVBoxLayout(central)

        # Barra superior
        top_bar = QHBoxLayout()
        self.btn_carregar = QPushButton("Carregar (pasta dados/)")
        self.btn_carregar.clicked.connect(self.on_carregar)

        self.btn_escolher = QPushButton("Escolher pasta...")
        self.btn_escolher.clicked.connect(self.on_escolher_pasta)

        self.btn_limpar = QPushButton("Limpar filtros")
        self.btn_limpar.clicked.connect(self.on_limpar)
        self.btn_limpar.setEnabled(False)

        self.btn_exportar = QPushButton("Exportar Excel v4")
        self.btn_exportar.clicked.connect(self.on_exportar)
        self.btn_exportar.setEnabled(False)

        self.lbl_status = QLabel("Status: aguardando...")
        self.lbl_status.setStyleSheet("font-weight:bold;")

        top_bar.addWidget(self.btn_carregar)
        top_bar.addWidget(self.btn_escolher)
        top_bar.addWidget(self.btn_limpar)
        top_bar.addWidget(self.btn_exportar)
        top_bar.addStretch()
        top_bar.addWidget(self.lbl_status)

        main_layout.addLayout(top_bar)

        # Filtros
        filtros_box = QGroupBox("Filtros")
        filtros_layout = QHBoxLayout(filtros_box)

        self.input_busca = QLineEdit()
        self.input_busca.setPlaceholderText("Buscar (produto, EAN, UF, canal, cliente, CFOP, CST...)")
        self.input_busca.textChanged.connect(self.aplicar_filtros)

        self.combo_uf = QComboBox()
        self.combo_uf.currentIndexChanged.connect(self.aplicar_filtros)

        self.combo_canal = QComboBox()
        self.combo_canal.currentIndexChanged.connect(self.aplicar_filtros)

        self.combo_categoria = QComboBox()
        self.combo_categoria.currentIndexChanged.connect(self.aplicar_filtros)

        self.combo_status = QComboBox()
        self.combo_status.currentIndexChanged.connect(self.aplicar_filtros)

        self.lbl_linhas = QLabel("Linhas: 0")
        self.lbl_linhas.setStyleSheet("font-weight:bold;")

        filtros_layout.addWidget(QLabel("Texto:"))
        filtros_layout.addWidget(self.input_busca, 2)
        filtros_layout.addWidget(self.combo_uf)
        filtros_layout.addWidget(self.combo_canal)
        filtros_layout.addWidget(self.combo_categoria)
        filtros_layout.addWidget(self.combo_status)
        filtros_layout.addWidget(self.lbl_linhas)

        main_layout.addWidget(filtros_box)

        # Tabs principais
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)

        # TAB 1 - Tabela
        self.tab_tabela = QWidget()
        self.tabs.addTab(self.tab_tabela, "Tabela")
        t1 = QVBoxLayout(self.tab_tabela)
        self.table_base = QTableView()
        self.table_base.setSortingEnabled(True)
        self.model_base = PandasTableModel(pd.DataFrame())
        self.table_base.setModel(self.model_base)
        t1.addWidget(self.table_base)

        # TAB 2 - KPIs
        self.tab_kpis = QWidget()
        self.tabs.addTab(self.tab_kpis, "KPIs")
        t2 = QHBoxLayout(self.tab_kpis)

        kpi_box = QGroupBox("KPIs (dinâmicos)")
        kpi_form = QFormLayout(kpi_box)

        self.lbl_fat_bruto = QLabel("-")
        self.lbl_fat_liq = QLabel("-")
        self.lbl_impacto_cancel = QLabel("-")
        self.lbl_ticket = QLabel("-")
        self.lbl_vendas = QLabel("-")
        self.lbl_itens = QLabel("-")
        self.lbl_margem = QLabel("-")

        kpi_form.addRow("Faturamento bruto:", self.lbl_fat_bruto)
        kpi_form.addRow("Faturamento líquido:", self.lbl_fat_liq)
        kpi_form.addRow("Impacto cancel/deneg:", self.lbl_impacto_cancel)
        kpi_form.addRow("Ticket médio:", self.lbl_ticket)
        kpi_form.addRow("Qtd vendas:", self.lbl_vendas)
        kpi_form.addRow("Qtd itens:", self.lbl_itens)
        kpi_form.addRow("Margem bruta:", self.lbl_margem)

        t2.addWidget(kpi_box, 1)

        # TAB 3 - Gráficos (sub-tabs)
        self.tab_graficos = QWidget()
        self.tabs.addTab(self.tab_graficos, "Gráficos")
        t3 = QVBoxLayout(self.tab_graficos)
        self.tabs_graf = QTabWidget()
        t3.addWidget(self.tabs_graf)

        # Gráfico: Top Produtos
        self.tab_top_prod = QWidget()
        self.tabs_graf.addTab(self.tab_top_prod, "Top Produtos")
        ltp = QVBoxLayout(self.tab_top_prod)
        self.fig_top = Figure(figsize=(6, 4))
        self.canvas_top = FigureCanvas(self.fig_top)
        ltp.addWidget(self.canvas_top)

        # Gráfico: Faturamento Mensal
        self.tab_mes = QWidget()
        self.tabs_graf.addTab(self.tab_mes, "Faturamento Mensal")
        lmes = QVBoxLayout(self.tab_mes)
        self.fig_mes = Figure(figsize=(6, 4))
        self.canvas_mes = FigureCanvas(self.fig_mes)
        lmes.addWidget(self.canvas_mes)

        # Gráfico: Por Canal
        self.tab_canal = QWidget()
        self.tabs_graf.addTab(self.tab_canal, "Por Canal")
        lcan = QVBoxLayout(self.tab_canal)
        self.fig_canal = Figure(figsize=(6, 4))
        self.canvas_canal = FigureCanvas(self.fig_canal)
        lcan.addWidget(self.canvas_canal)

        # Gráfico: Margem por Categoria
        self.tab_margem_cat = QWidget()
        self.tabs_graf.addTab(self.tab_margem_cat, "Margem por Categoria")
        lmc = QVBoxLayout(self.tab_margem_cat)
        self.fig_margem = Figure(figsize=(6, 4))
        self.canvas_margem = FigureCanvas(self.fig_margem)
        lmc.addWidget(self.canvas_margem)

        # TAB 4 - Clientes
        self.tab_clientes = QWidget()
        self.tabs.addTab(self.tab_clientes, "Clientes")
        t4 = QVBoxLayout(self.tab_clientes)

        self.lbl_clientes = QLabel("Top Clientes (por faturamento)")
        self.lbl_clientes.setStyleSheet("font-weight:bold;")
        t4.addWidget(self.lbl_clientes)

        self.table_cli = QTableView()
        self.table_cli.setSortingEnabled(True)
        self.model_cli = PandasTableModel(pd.DataFrame())
        self.table_cli.setModel(self.model_cli)
        t4.addWidget(self.table_cli)

        self.fig_cli = Figure(figsize=(6, 3))
        self.canvas_cli = FigureCanvas(self.fig_cli)
        t4.addWidget(self.canvas_cli)

        # TAB 5 - Fiscal
        self.tab_fiscal = QWidget()
        self.tabs.addTab(self.tab_fiscal, "Auditoria Fiscal")
        t5 = QVBoxLayout(self.tab_fiscal)

        self.lbl_fiscal = QLabel("ICMS/PIS/COFINS por UF + CFOP (se existirem colunas no dataset)")
        self.lbl_fiscal.setStyleSheet("font-weight:bold;")
        t5.addWidget(self.lbl_fiscal)

        self.table_fiscal = QTableView()
        self.table_fiscal.setSortingEnabled(True)
        self.model_fiscal = PandasTableModel(pd.DataFrame())
        self.table_fiscal.setModel(self.model_fiscal)
        t5.addWidget(self.table_fiscal)

        self.fig_fiscal = Figure(figsize=(6, 3))
        self.canvas_fiscal = FigureCanvas(self.fig_fiscal)
        t5.addWidget(self.canvas_fiscal)

        # TAB 6 - Qualidade
        self.tab_quality = QWidget()
        self.tabs.addTab(self.tab_quality, "Qualidade de Dados")
        t6 = QVBoxLayout(self.tab_quality)
        self.lbl_quality = QLabel("Carregue os dados para ver a auditoria de qualidade.")
        self.lbl_quality.setTextInteractionFlags(Qt.TextSelectableByMouse)
        t6.addWidget(self.lbl_quality)

        # Auto-load
        if self.data_dir.exists():
            self.on_carregar()

    # =========================================================
    # Carregar / preparar
    # =========================================================
    def on_escolher_pasta(self):
        folder = QFileDialog.getExistingDirectory(self, "Selecione a pasta dos CSVs")
        if folder:
            self.data_dir = Path(folder)
            self.on_carregar()

    def on_carregar(self):
        try:
            faltando = checar_arquivos_erp(self.data_dir)
            if faltando:
                raise FileNotFoundError(f"Arquivos faltando em {self.data_dir}: {faltando}")

            self.base = carregar_base_erp(self.data_dir)
            self.base_filtrada = self.base.copy()

            self.preencher_combos()

            self.atualizar_tudo()

            self.btn_limpar.setEnabled(True)
            self.btn_exportar.setEnabled(True)
            self.lbl_status.setText(f"Status: OK | Base: {self.base.shape[0]} linhas / {self.base.shape[1]} colunas")

        except Exception as e:
            QMessageBox.critical(self, "Erro ao carregar", str(e))
            self.lbl_status.setText("Status: ERRO ao carregar")

    def on_limpar(self):
        self.input_busca.setText("")
        self.combo_uf.setCurrentIndex(0)
        self.combo_canal.setCurrentIndex(0)
        self.combo_categoria.setCurrentIndex(0)
        self.combo_status.setCurrentIndex(0)

        self.base_filtrada = self.base.copy()
        self.atualizar_tudo()

    def preencher_combos(self):
        # bloquear sinais
        self.combo_uf.blockSignals(True)
        self.combo_canal.blockSignals(True)
        self.combo_categoria.blockSignals(True)
        self.combo_status.blockSignals(True)

        self.combo_uf.clear()
        self.combo_canal.clear()
        self.combo_categoria.clear()
        self.combo_status.clear()

        self.combo_uf.addItem("UF: Todas")
        self.combo_canal.addItem("Canal: Todos")
        self.combo_categoria.addItem("Categoria: Todas")
        self.combo_status.addItem("Status: Todos")

        if "uf_destino" in self.base.columns:
            for u in sorted(self.base["uf_destino"].dropna().unique().tolist()):
                self.combo_uf.addItem(f"UF: {u}")

        if "canal" in self.base.columns:
            for c in sorted(self.base["canal"].dropna().unique().tolist()):
                self.combo_canal.addItem(f"Canal: {c}")

        if "categoria" in self.base.columns:
            for cat in sorted(self.base["categoria"].dropna().unique().tolist()):
                self.combo_categoria.addItem(f"Categoria: {cat}")

        if "status" in self.base.columns:
            for st in sorted(self.base["status"].dropna().unique().tolist()):
                self.combo_status.addItem(f"Status: {st}")

        # desbloquear
        self.combo_uf.blockSignals(False)
        self.combo_canal.blockSignals(False)
        self.combo_categoria.blockSignals(False)
        self.combo_status.blockSignals(False)

    # =========================================================
    # Filtragem
    # =========================================================
    def aplicar_filtros(self):
        if self.base.empty:
            return

        df = self.base.copy()

        # Texto
        termo = self.input_busca.text().strip().lower()
        if termo:
            cols_preferidas = [c for c in df.columns if c.lower() in [
                "produto", "ean", "uf_destino", "canal", "categoria",
                "cfop", "cst_icms", "nome", "marca"
            ]]
            if not cols_preferidas:
                cols_preferidas = df.columns.tolist()

            mask = df[cols_preferidas].astype(str).apply(
                lambda col: col.str.lower().str.contains(termo, na=False)
            ).any(axis=1)
            df = df[mask]

        # UF
        uf_sel = self.combo_uf.currentText().replace("UF: ", "")
        if uf_sel != "Todas" and "uf_destino" in df.columns:
            df = df[df["uf_destino"] == uf_sel]

        # Canal
        canal_sel = self.combo_canal.currentText().replace("Canal: ", "")
        if canal_sel != "Todos" and "canal" in df.columns:
            df = df[df["canal"] == canal_sel]

        # Categoria
        cat_sel = self.combo_categoria.currentText().replace("Categoria: ", "")
        if cat_sel != "Todas" and "categoria" in df.columns:
            df = df[df["categoria"] == cat_sel]

        # Status
        st_sel = self.combo_status.currentText().replace("Status: ", "")
        if st_sel != "Todos" and "status" in df.columns:
            df = df[df["status"] == st_sel]

        self.base_filtrada = df.copy()
        self.atualizar_tudo()

    # =========================================================
    # KPIs e agregações
    # =========================================================
    def calcular_faturamento_liquido(self, df: pd.DataFrame) -> tuple[float, float, float]:
        """
        Retorna:
        - faturamento bruto
        - faturamento líquido (exclui canceladas/denegadas se existir status)
        - impacto (bruto - líquido)
        """
        base = df.copy()
        if "valor_total" not in base.columns:
            return 0.0, 0.0, 0.0

        fat_bruto = float(base["valor_total"].sum())

        if "status" not in base.columns:
            return fat_bruto, fat_bruto, 0.0

        status = base["status"].astype(str).str.upper()
        invalidos = status.str.contains("CANCEL") | status.str.contains("DENEG") | status.str.contains("INUTIL")
        base_ok = base[~invalidos].copy()

        fat_liq = float(base_ok["valor_total"].sum())
        impacto = fat_bruto - fat_liq
        return fat_bruto, fat_liq, impacto

    def calcular_kpis(self, df: pd.DataFrame) -> dict:
        base = df.copy()
        if "valor_total" not in base.columns:
            return {}

        fat_bruto, fat_liq, impacto = self.calcular_faturamento_liquido(base)

        qtd_vendas = int(base["venda_id"].nunique()) if "venda_id" in base.columns else int(base.shape[0])
        ticket = float(base.groupby("venda_id")["valor_total"].sum().mean()) if "venda_id" in base.columns else float(fat_bruto)

        qtd_itens = int(base.shape[0])

        margem_pct = None
        if "custo_total" in base.columns:
            custo = float(base["custo_total"].sum())
            lucro = fat_bruto - custo
            margem_pct = float(lucro / fat_bruto) if fat_bruto != 0 else 0.0

        return {
            "fat_bruto": fat_bruto,
            "fat_liq": fat_liq,
            "impacto": impacto,
            "qtd_vendas": qtd_vendas,
            "ticket": ticket,
            "qtd_itens": qtd_itens,
            "margem_pct": margem_pct
        }

    def top_produtos(self, df: pd.DataFrame) -> pd.Series:
        base = df.copy()
        if "produto" not in base.columns or "valor_total" not in base.columns:
            return pd.Series(dtype=float)
        return (base.groupby("produto")["valor_total"]
                .sum()
                .sort_values(ascending=False)
                .head(10))

    def top_clientes(self, df: pd.DataFrame) -> pd.DataFrame:
        base = df.copy()
        if "nome" not in base.columns or "valor_total" not in base.columns:
            return pd.DataFrame()

        top = (base.groupby("nome")["valor_total"]
               .sum()
               .sort_values(ascending=False)
               .head(10)
               .reset_index())

        top.rename(columns={"valor_total": "faturamento"}, inplace=True)
        return top

    def serie_mensal(self, df: pd.DataFrame) -> pd.Series:
        base = df.copy()
        if "data" not in base.columns or "valor_total" not in base.columns:
            return pd.Series(dtype=float)

        base = base.dropna(subset=["data"])
        if base.empty:
            return pd.Series(dtype=float)

        serie = (base.groupby(base["data"].dt.to_period("M"))["valor_total"]
                 .sum()
                 .sort_index())
        return serie

    def por_canal(self, df: pd.DataFrame) -> pd.Series:
        base = df.copy()
        if "canal" not in base.columns or "valor_total" not in base.columns:
            return pd.Series(dtype=float)

        return (base.groupby("canal")["valor_total"]
                .sum()
                .sort_values(ascending=False))

    def margem_por_categoria(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Necessita: categoria, valor_total e custo_total
        Retorna tabela com margem por categoria.
        """
        base = df.copy()
        if not all(c in base.columns for c in ["categoria", "valor_total", "custo_total"]):
            return pd.DataFrame()

        agg = (base.groupby("categoria")[["valor_total", "custo_total"]]
               .sum()
               .reset_index())

        agg["lucro_bruto"] = agg["valor_total"] - agg["custo_total"]
        agg["margem_pct"] = agg["lucro_bruto"] / agg["valor_total"].replace({0: 1})

        agg.rename(columns={"valor_total": "faturamento", "custo_total": "custo"}, inplace=True)
        agg = agg.sort_values("margem_pct", ascending=False)
        return agg

    def auditoria_fiscal(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Requer colunas fiscais:
        uf_destino, cfop, valor_icms, valor_pis, valor_cofins
        """
        base = df.copy()
        cols = ["uf_destino", "cfop", "valor_icms", "valor_pis", "valor_cofins"]

        if not all(c in base.columns for c in cols):
            return pd.DataFrame()

        fiscal = (base.groupby(["uf_destino", "cfop"])[["valor_icms", "valor_pis", "valor_cofins", "valor_total"]]
                  .sum()
                  .reset_index())

        fiscal.rename(columns={
            "valor_total": "faturamento",
            "valor_icms": "icms",
            "valor_pis": "pis",
            "valor_cofins": "cofins"
        }, inplace=True)

        fiscal = fiscal.sort_values("faturamento", ascending=False).head(50)
        return fiscal

    # =========================================================
    # Qualidade de dados
    # =========================================================
    def quality_text(self, df: pd.DataFrame) -> str:
        base = df.copy()
        linhas = []
        linhas.append(f"- Linhas: {base.shape[0]}")
        linhas.append(f"- Colunas: {base.shape[1]}")

        # Nulos
        nulos = base.isna().sum().sort_values(ascending=False).head(15)
        linhas.append("\nTop 15 colunas com nulos:")
        for col, qtd in nulos.items():
            if qtd > 0:
                linhas.append(f"  • {col}: {int(qtd)}")

        # EAN inválido
        if "ean" in base.columns:
            ean_len = base["ean"].astype(str).str.len()
            inv = int((ean_len != 13).sum())
            linhas.append(f"\nEAN inválidos (len != 13): {inv}")

            dup = base[base.duplicated(subset=["ean"], keep=False)]
            linhas.append(f"EAN duplicados (distintos): {dup['ean'].nunique()}")

        # Outliers preco_unitario (IQR)
        if "preco_unitario" in base.columns:
            q1 = base["preco_unitario"].quantile(0.25)
            q3 = base["preco_unitario"].quantile(0.75)
            iqr = q3 - q1
            lim_sup = q3 + 1.5 * iqr
            out = int((base["preco_unitario"] > lim_sup).sum())
            linhas.append(f"\nOutliers de preco_unitario (IQR): {out}")

        return "🧪 Qualidade de Dados (recorte atual)\n\n" + "\n".join(linhas)

    # =========================================================
    # Atualização UI
    # =========================================================
    def atualizar_tudo(self):
        df = self.base_filtrada if not self.base_filtrada.empty else self.base

        # tabela base
        self.model_base.update(df)
        self.lbl_linhas.setText(f"Linhas: {len(df):,}".replace(",", "."))

        # KPIs
        self.atualizar_kpis(df)

        # gráficos
        self.atualizar_graficos(df)

        # clientes
        self.atualizar_clientes(df)

        # fiscal
        self.atualizar_fiscal(df)

        # qualidade
        self.lbl_quality.setText(self.quality_text(df))

    def atualizar_kpis(self, df: pd.DataFrame):
        k = self.calcular_kpis(df)
        if not k:
            self.lbl_fat_bruto.setText("—")
            self.lbl_fat_liq.setText("—")
            self.lbl_impacto_cancel.setText("—")
            self.lbl_ticket.setText("—")
            self.lbl_vendas.setText("—")
            self.lbl_itens.setText("—")
            self.lbl_margem.setText("—")
            return

        self.lbl_fat_bruto.setText(format_brl(k["fat_bruto"]))
        self.lbl_fat_liq.setText(format_brl(k["fat_liq"]))
        self.lbl_impacto_cancel.setText(format_brl(k["impacto"]))
        self.lbl_ticket.setText(format_brl(k["ticket"]))
        self.lbl_vendas.setText(str(k["qtd_vendas"]))
        self.lbl_itens.setText(str(k["qtd_itens"]))

        if k["margem_pct"] is None:
            self.lbl_margem.setText("—")
        else:
            self.lbl_margem.setText(f"{k['margem_pct']*100:.2f}%".replace(".", ","))

    def atualizar_graficos(self, df: pd.DataFrame):
        # Top Produtos
        top = self.top_produtos(df)
        self.fig_top.clear()
        ax = self.fig_top.add_subplot(111)
        ax.set_title("Top 10 Produtos por Faturamento")
        if len(top) > 0:
            ax.bar(top.index.astype(str), top.values)
            ax.tick_params(axis="x", rotation=45)
        else:
            ax.text(0.5, 0.5, "Coluna 'produto' não disponível.", ha="center", va="center")
        self.fig_top.tight_layout()
        self.canvas_top.draw()

        # Série Mensal
        serie = self.serie_mensal(df)
        self.fig_mes.clear()
        ax2 = self.fig_mes.add_subplot(111)
        ax2.set_title("Faturamento Mensal")
        if len(serie) > 0:
            x = [str(p) for p in serie.index]
            ax2.plot(x, serie.values, marker="o")
            ax2.tick_params(axis="x", rotation=45)
        else:
            ax2.text(0.5, 0.5, "Coluna 'data' não disponível.", ha="center", va="center")
        self.fig_mes.tight_layout()
        self.canvas_mes.draw()

        # Por Canal
        canal = self.por_canal(df)
        self.fig_canal.clear()
        ax3 = self.fig_canal.add_subplot(111)
        ax3.set_title("Faturamento por Canal")
        if len(canal) > 0:
            ax3.bar(canal.index.astype(str), canal.values)
        else:
            ax3.text(0.5, 0.5, "Coluna 'canal' não disponível.", ha="center", va="center")
        self.fig_canal.tight_layout()
        self.canvas_canal.draw()

        # Margem por Categoria
        marg = self.margem_por_categoria(df)
        self.fig_margem.clear()
        ax4 = self.fig_margem.add_subplot(111)
        ax4.set_title("Margem Bruta (%) por Categoria")
        if len(marg) > 0:
            topm = marg.head(10).copy()
            ax4.bar(topm["categoria"].astype(str), topm["margem_pct"] * 100)
            ax4.tick_params(axis="x", rotation=45)
        else:
            ax4.text(0.5, 0.5, "Necessário: categoria + custo_total", ha="center", va="center")
        self.fig_margem.tight_layout()
        self.canvas_margem.draw()

    def atualizar_clientes(self, df: pd.DataFrame):
        top_cli = self.top_clientes(df)
        self.model_cli.update(top_cli)

        # gráfico
        self.fig_cli.clear()
        ax = self.fig_cli.add_subplot(111)
        ax.set_title("Top 10 Clientes por Faturamento")
        if len(top_cli) > 0:
            ax.bar(top_cli["nome"].astype(str), top_cli["faturamento"].values)
            ax.tick_params(axis="x", rotation=45)
        else:
            ax.text(0.5, 0.5, "Coluna 'nome' não disponível.", ha="center", va="center")
        self.fig_cli.tight_layout()
        self.canvas_cli.draw()

    def atualizar_fiscal(self, df: pd.DataFrame):
        fiscal = self.auditoria_fiscal(df)
        self.model_fiscal.update(fiscal)

        # gráfico ICMS por UF (somente se fiscal tiver dados)
        self.fig_fiscal.clear()
        ax = self.fig_fiscal.add_subplot(111)
        ax.set_title("ICMS por UF (top 10)")

        if len(fiscal) > 0:
            icms_uf = fiscal.groupby("uf_destino")["icms"].sum().sort_values(ascending=False).head(10)
            ax.bar(icms_uf.index.astype(str), icms_uf.values)
        else:
            ax.text(0.5, 0.5, "Sem colunas fiscais (icms/pis/cofins/cfop).", ha="center", va="center")

        self.fig_fiscal.tight_layout()
        self.canvas_fiscal.draw()

    # =========================================================
    # Export Excel v4
    # =========================================================
    def on_exportar(self):
        if self.base.empty:
            return

        df = self.base_filtrada if not self.base_filtrada.empty else self.base

        arquivo, _ = QFileDialog.getSaveFileName(
            self,
            "Salvar Excel v4",
            str(self.base_dir / "relatorio_dashboard_v4.xlsx"),
            "Excel (*.xlsx)"
        )
        if not arquivo:
            return

        try:
            kpis = self.calcular_kpis(df)
            top_prod = self.top_produtos(df)
            top_cli = self.top_clientes(df)
            serie = self.serie_mensal(df)
            canal = self.por_canal(df)
            marg = self.margem_por_categoria(df)
            fiscal = self.auditoria_fiscal(df)

            df_kpis = pd.DataFrame([{
                "faturamento_bruto": kpis.get("fat_bruto", 0.0),
                "faturamento_liquido": kpis.get("fat_liq", 0.0),
                "impacto_cancel_deneg": kpis.get("impacto", 0.0),
                "ticket_medio": kpis.get("ticket", 0.0),
                "qtd_vendas": kpis.get("qtd_vendas", 0),
                "qtd_itens": kpis.get("qtd_itens", 0),
                "margem_bruta_pct": (kpis["margem_pct"] if kpis.get("margem_pct") is not None else None)
            }])

            df_top_prod = top_prod.reset_index()
            if len(df_top_prod) > 0:
                df_top_prod.columns = ["produto", "faturamento"]

            df_serie = serie.reset_index()
            if len(df_serie) > 0:
                df_serie.columns = ["mes", "faturamento"]

            df_canal = canal.reset_index()
            if len(df_canal) > 0:
                df_canal.columns = ["canal", "faturamento"]

            df_quality = pd.DataFrame({"texto": [self.quality_text(df)]})

            sheets = {
                "Base_Filtrada": df,
                "KPIs": df_kpis,
                "Top_Produtos": df_top_prod,
                "Top_Clientes": top_cli,
                "Faturamento_Mensal": df_serie,
                "Faturamento_Canal": df_canal,
                "Margem_Categoria": marg,
                "Fiscal_UF_CFOP": fiscal,
                "Qualidade": df_quality
            }

            export_excel_formatado(arquivo, sheets)

            QMessageBox.information(self, "Exportado", f"✅ Excel v4 salvo:\n{arquivo}")

        except Exception as e:
            QMessageBox.critical(self, "Erro ao exportar", str(e))


def main():
    app = QApplication(sys.argv)
    w = DashboardV4()
    w.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
